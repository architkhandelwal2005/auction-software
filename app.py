from flask import Flask, render_template, request, jsonify, send_from_directory, session, redirect, url_for, g
import sqlite3, os, webbrowser, csv, io, uuid, json, re, threading
import urllib.request as _ureq
from threading import Timer

# ── Load .env (DATABASE_URL, etc.) ──────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
except ImportError:
    pass

DATABASE_URL = os.environ.get('DATABASE_URL', '')
USE_PG = bool(DATABASE_URL)

SHEETS_WEBHOOK = os.environ.get(
    'SHEETS_WEBHOOK',
    'https://script.google.com/macros/s/AKfycbwDEpBKxUPlZG5mT1ea8w-C3csMiJT13m1ofK9n9SNoy5DzkBoVfURocIjkjGSVbek/exec'
)

# ── PostgreSQL thin wrapper (speaks the same API as sqlite3) ─────────────────
def _pg_sql(sql):
    """Convert SQLite SQL dialect → PostgreSQL."""
    sql = sql.replace('?', '%s')
    def _upsert(m):
        tbl, cols_raw, vals = m.group(1), m.group(2), m.group(3)
        cols = [c.strip() for c in cols_raw.split(',')]
        updates = ', '.join(f'{c}=EXCLUDED.{c}' for c in cols[1:])
        tail = f'UPDATE SET {updates}' if updates else 'NOTHING'
        return f'INSERT INTO {tbl} ({cols_raw}) VALUES ({vals}) ON CONFLICT ({cols[0]}) DO {tail}'
    sql = re.sub(
        r'INSERT\s+OR\s+REPLACE\s+INTO\s+(\w+)\s*\(([^)]+)\)\s*VALUES\s*\(([^)]+)\)',
        _upsert, sql, flags=re.IGNORECASE)
    # Double-quoted string literals → single-quoted (SQLite allows both; PG does not)
    sql = re.sub(r'"([^"]*)"', r"'\1'", sql)
    return sql

class _PGCur:
    def __init__(self, raw):
        self._c = raw
        self.lastrowid = None

    def execute(self, sql, params=()):
        converted = _pg_sql(sql)
        self._c.execute(converted, params or ())
        return self

    def fetchone(self):  return self._c.fetchone()
    def fetchall(self):  return self._c.fetchall()
    def __iter__(self):  return iter(self.fetchall())
    def __getitem__(self, k): return self.fetchall()[k]

class _PGConn:
    def __init__(self, raw):
        self._r = raw
        self._closed = False

    def cursor(self):
        import psycopg2.extras
        return _PGCur(self._r.cursor(cursor_factory=psycopg2.extras.RealDictCursor))

    def execute(self, sql, params=()):
        c = self.cursor(); c.execute(sql, params); return c

    def commit(self):   self._r.commit()
    def rollback(self): self._r.rollback()
    def close(self):
        # No-op: pooled connections are returned to the pool by the request
        # teardown (close_db), never closed by endpoint code. Closing here would
        # drain the pool one connection at a time.
        self._closed = True

# ── Google Sheets async push ─────────────────────────────────────────────────
def _sheets_push(payload):
    try:
        data = json.dumps(payload).encode()
        req = _ureq.Request(SHEETS_WEBHOOK, data=data,
                            headers={'Content-Type': 'application/json'}, method='POST')
        _ureq.urlopen(req, timeout=10)
    except Exception as e:
        print(f'[Sheets] sync error: {e}', flush=True)

def sheets_push_async(payload):
    threading.Thread(target=_sheets_push, args=(payload,), daemon=True).start()

# ── Local Excel backup ──────────────────────────────────────────────────────
# Whenever running locally (no cloud DB, or even with one — this is a belt-
# and-braces backup), rewrite a full snapshot of the auction to a local .xlsx
# after every sale/undo/edit/reset. Runs in a background thread with its own
# DB connection so it never blocks or breaks the actual auction action, and
# a failure here is only logged, never surfaced to the user.
EXCEL_BACKUP_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'auction_backup.xlsx')

def archive_excel_backup(reason='archive'):
    """Preserve the current backup under a timestamped name before a reset or
    wipe. Without this, resetting an auction overwrites the only record of a
    completed one with an empty snapshot."""
    try:
        if not os.path.exists(EXCEL_BACKUP_PATH):
            return
        import shutil, datetime
        stamp = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
        dest = EXCEL_BACKUP_PATH.replace('.xlsx', f'-{reason}-{stamp}.xlsx')
        shutil.copy2(EXCEL_BACKUP_PATH, dest)
        print(f'[Excel Backup] archived to {dest}', flush=True)
    except Exception as e:
        print(f'[Excel Backup] archive failed: {e}', flush=True)

def _write_excel_backup():
    try:
        import openpyxl
        if USE_PG:
            import psycopg2, psycopg2.extras
            raw = psycopg2.connect(DATABASE_URL)
            cur = raw.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        else:
            raw = sqlite3.connect(DB_FILE, timeout=20)
            raw.row_factory = sqlite3.Row
            cur = raw.cursor()

        cur.execute('SELECT * FROM teams ORDER BY id')
        teams = [dict(r) for r in cur.fetchall()]
        cur.execute('SELECT * FROM players ORDER BY id')
        players = [dict(r) for r in cur.fetchall()]
        raw.close()
        team_by_id = {t['id']: t for t in teams}

        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = 'Sold Players'
        ws1.append(['Name', 'Category', 'Base Price', 'Sold Price', 'Team', 'Sold At'])
        sold = [p for p in players if p.get('status') == 'sold']
        for p in sorted(sold, key=lambda x: str(x.get('sold_at') or '')):
            ws1.append([p.get('name'), p.get('category') or '', p.get('base_price') or 0,
                        p.get('sold_price') or 0, team_by_id.get(p.get('team_id'), {}).get('name', ''),
                        str(p.get('sold_at') or '')])

        ws2 = wb.create_sheet('All Players')
        ws2.append(['Name', 'Category', 'Base Price', 'Status', 'Team', 'Sold Price'])
        for p in players:
            ws2.append([p.get('name'), p.get('category') or '', p.get('base_price') or 0,
                        p.get('status') or '', team_by_id.get(p.get('team_id'), {}).get('name', ''),
                        p.get('sold_price') or ''])

        ws3 = wb.create_sheet('Teams')
        ws3.append(['Team', 'Total Budget', 'Remaining Budget', 'Spent', 'Players Bought'])
        for t in teams:
            spent = (t.get('total_budget') or 0) - (t.get('remaining_budget') or 0)
            count = sum(1 for p in players if p.get('team_id') == t['id'] and p.get('status') == 'sold')
            ws3.append([t.get('name'), t.get('total_budget') or 0, t.get('remaining_budget') or 0, spent, count])

        for ws in (ws1, ws2, ws3):
            for col_cells in ws.columns:
                length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(40, max(10, length + 2))

        wb.save(EXCEL_BACKUP_PATH)
    except Exception as e:
        print(f'[Excel Backup] failed: {e}', flush=True)

def excel_backup_async():
    threading.Thread(target=_write_excel_backup, daemon=True).start()

app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = 'super_secret_auction_pro_key_2026'
DB_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'auction.db')
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Cache-busting version for front-end assets. Uses the newest mtime of the JSX
# bundles so every deploy serves a fresh URL and browsers never run stale JS
# (e.g. an old "Start Over" without the wipe-password prompt).
def _asset_version():
    base = os.path.dirname(os.path.abspath(__file__))
    latest = 0.0
    for rel in ('static/app.jsx', 'static/team_app.jsx', 'static/auction_display.js', 'static/auction_display.css'):
        try:
            latest = max(latest, os.path.getmtime(os.path.join(base, rel)))
        except OSError:
            pass
    return str(int(latest))

@app.context_processor
def inject_asset_version():
    return {'asset_v': _asset_version()}

# ── PostgreSQL connection pool ──
# Reuse warm connections instead of opening a fresh one per request (a fresh
# connect to a remote DB costs ~0.5s). This is the single biggest speed win for
# the deployed/cloud app. Each gunicorn worker gets its own small pool.
_pg_pool = None
def _get_pg_pool():
    global _pg_pool
    if _pg_pool is None:
        import psycopg2.pool
        _pg_pool = psycopg2.pool.ThreadedConnectionPool(1, 10, DATABASE_URL)
    return _pg_pool

def get_db():
    if 'db' not in g:
        if USE_PG:
            raw = _get_pg_pool().getconn()
            g.db = _PGConn(raw)
        else:
            g.db = sqlite3.connect(DB_FILE, timeout=20, check_same_thread=False)
            g.db.row_factory = sqlite3.Row
            g.db.execute("PRAGMA foreign_keys = ON")
    return g.db

@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop('db', None)
    if db is None:
        return
    if USE_PG:
        # Always reset transaction state before returning to the pool, so a
        # connection is never handed back "idle in transaction" (which would
        # hold locks and stall the next request). Then return it to the pool
        # instead of closing — this is what makes reuse fast.
        try: db._r.rollback()
        except Exception: pass
        try: _get_pg_pool().putconn(db._r)
        except Exception:
            try: db._r.close()
            except Exception: pass
    else:
        if exception:
            try: db.rollback()
            except Exception: pass
        try: db.close()
        except Exception: pass

def init_db():
    if USE_PG:
        _init_db_pg()
    else:
        _init_db_sqlite()

def _init_db_sqlite():
    conn = sqlite3.connect(DB_FILE, timeout=20)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL")
    except Exception:
        pass
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS category_rules (
        id INTEGER PRIMARY KEY AUTOINCREMENT, category TEXT NOT NULL,
        base_price REAL DEFAULT 0, min_per_team INTEGER DEFAULT 0, max_per_team INTEGER DEFAULT 99)''')
    c.execute('''CREATE TABLE IF NOT EXISTS teams (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
        total_budget REAL NOT NULL, remaining_budget REAL NOT NULL,
        color TEXT DEFAULT '#3b82f6', logo_url TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS players (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, category TEXT,
        base_price REAL DEFAULT 0, status TEXT DEFAULT 'unsold', team_id INTEGER,
        sold_price REAL, photo_url TEXT, sold_at TIMESTAMP,
        FOREIGN KEY (team_id) REFERENCES teams (id))''')
    c.execute('''CREATE TABLE IF NOT EXISTS auction_state (key TEXT PRIMARY KEY, value TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS action_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT, action_type TEXT NOT NULL,
        player_id INTEGER NOT NULL, player_name TEXT, old_team_id INTEGER, new_team_id INTEGER,
        old_sold_price REAL, new_sold_price REAL, old_status TEXT, new_status TEXT,
        base_price REAL, category TEXT, photo_url TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    team_cols = [r[1] for r in c.execute("PRAGMA table_info(teams)").fetchall()]
    if 'logo_url' not in team_cols:
        c.execute("ALTER TABLE teams ADD COLUMN logo_url TEXT")
    player_cols = [r[1] for r in c.execute("PRAGMA table_info(players)").fetchall()]
    for col, typ in [('photo_url', 'TEXT'), ('sold_at', 'TIMESTAMP'), ('attributes', 'TEXT')]:
        if col not in player_cols:
            c.execute(f"ALTER TABLE players ADD COLUMN {col} {typ}")
    cfg_count = c.execute("SELECT COUNT(*) as c FROM config").fetchone()['c']
    if cfg_count == 0:
        for k, v in {'event_name': 'Premier Auction 2026', 'common_base_price': '10',
                     'min_players_per_team': '10', 'bid_increment': '2.5', 'setup_done': 'false'}.items():
            c.execute("INSERT OR REPLACE INTO config (key, value) VALUES (?, ?)", (k, v))
    conn.commit()
    conn.close()

def _init_db_pg():
    import psycopg2
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)")
    cur.execute("""CREATE TABLE IF NOT EXISTS category_rules (
        id SERIAL PRIMARY KEY, category TEXT NOT NULL,
        base_price DOUBLE PRECISION DEFAULT 0,
        min_per_team INTEGER DEFAULT 0, max_per_team INTEGER DEFAULT 99)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS teams (
        id SERIAL PRIMARY KEY, name TEXT NOT NULL,
        total_budget DOUBLE PRECISION NOT NULL, remaining_budget DOUBLE PRECISION NOT NULL,
        color TEXT DEFAULT '#3b82f6', logo_url TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS players (
        id SERIAL PRIMARY KEY, name TEXT NOT NULL, category TEXT,
        base_price DOUBLE PRECISION DEFAULT 0, status TEXT DEFAULT 'unsold',
        team_id INTEGER REFERENCES teams(id), sold_price DOUBLE PRECISION,
        photo_url TEXT, sold_at TIMESTAMP, attributes TEXT)""")
    cur.execute("CREATE TABLE IF NOT EXISTS auction_state (key TEXT PRIMARY KEY, value TEXT)")
    cur.execute("""CREATE TABLE IF NOT EXISTS action_history (
        id SERIAL PRIMARY KEY, action_type TEXT NOT NULL, player_id INTEGER NOT NULL,
        player_name TEXT, old_team_id INTEGER, new_team_id INTEGER,
        old_sold_price DOUBLE PRECISION, new_sold_price DOUBLE PRECISION,
        old_status TEXT, new_status TEXT, base_price DOUBLE PRECISION,
        category TEXT, photo_url TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    # Idempotent column additions
    for stmt in [
        "ALTER TABLE teams ADD COLUMN IF NOT EXISTS logo_url TEXT",
        "ALTER TABLE players ADD COLUMN IF NOT EXISTS photo_url TEXT",
        "ALTER TABLE players ADD COLUMN IF NOT EXISTS sold_at TIMESTAMP",
        "ALTER TABLE players ADD COLUMN IF NOT EXISTS attributes TEXT",
    ]:
        cur.execute(stmt)
    # Seed config
    cur.execute("SELECT COUNT(*) FROM config")
    if cur.fetchone()[0] == 0:
        for k, v in {'event_name': 'Premier Auction 2026', 'common_base_price': '10',
                     'min_players_per_team': '10', 'bid_increment': '2.5', 'setup_done': 'false'}.items():
            cur.execute("INSERT INTO config (key, value) VALUES (%s, %s) ON CONFLICT (key) DO NOTHING", (k, v))
    conn.commit()
    conn.close()
    print('[DB] PostgreSQL schema ready.', flush=True)

# Initialize DB on load
init_db()

# ─── Serve uploads ───
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# ─── Authentication ───

@app.route('/log_error', methods=['POST'])
def log_error():
    print('BROWSER ERROR:', request.json, flush=True)
    return jsonify({'status': 'ok'})

@app.route('/login')
@app.route('/')
def login_portal():
    return render_template('login.html')

@app.route('/admin')
def admin_dashboard():
    if session.get('role') != 'admin':
        return redirect('/login')
    return render_template('index.html')

@app.route('/team/<int:team_id>')
def team_dashboard(team_id):
    if session.get('role') not in ['admin', 'team']:
        return redirect('/login')
    return render_template('team_view.html', team_id=team_id)

@app.route('/live')
@app.route('/view')
def live_view():
    return render_template('live_view.html')

@app.route('/presentation')
@app.route('/stage')
def presentation_view():
    return render_template('presentation.html')

@app.route('/cricket_auction')
def cricket_auction_view():
    return render_template('cricket_auction.html')

@app.route('/roster')
def roster_view():
    """Themed player-roster confirmation page. Also reachable from the setup
    wizard's review step; kept as a standalone URL for reprinting later."""
    return render_template('roster.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    data = request.json or {}
    role = data.get('role')
    password = data.get('password', '').strip()
    
    if role == 'admin':
        if password == 'admin@123':
            session['user'] = 'admin'
            session['role'] = 'admin'
            return jsonify({'success': True, 'url': '/admin'})
        return jsonify({'error': 'Incorrect admin password'}), 401
        
    elif role == 'team':
        team_id = data.get('team_id')
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT name FROM teams WHERE id=?', (team_id,))
        row = c.fetchone()
        conn.close()
        
        if not row:
            return jsonify({'error': 'Team not found'}), 404
            
        team_name = row['name']
        expected_pass = team_name.replace(' ', '').lower()
        entered_pass = password.replace(' ', '').lower()
        
        if expected_pass == entered_pass:
            session['user'] = team_name
            session['role'] = 'team'
            session['team_id'] = team_id
            return jsonify({'success': True, 'url': f'/team/{team_id}'})
        else:
            return jsonify({'error': 'Incorrect password'}), 401
            
    elif role == 'spectator':
        session['user'] = 'spectator'
        session['role'] = 'viewer'
        return jsonify({'success': True, 'url': '/live'})
        
    return jsonify({'error': 'Invalid role'}), 400

@app.route('/api/auth/me')
def auth_me():
    return jsonify({
        'logged_in': bool(session.get('user')),
        'username': session.get('user'),
        'role': session.get('role'),
        'team_id': session.get('team_id')
    })


@app.route('/report')
def report_view():
    if not session.get('role'):
        return redirect('/login')
    return render_template('report.html')

# ─── Config / Setup ───
@app.route('/api/config', methods=['GET'])
def get_config():
    conn = get_db()
    rows = conn.execute('SELECT * FROM config').fetchall()
    config = {r['key']: r['value'] for r in rows}
    rules = conn.execute('SELECT * FROM category_rules').fetchall()
    conn.close()
    return jsonify({'config': config, 'category_rules': [dict(r) for r in rules]})

@app.route('/api/config', methods=['POST'])
def save_config():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    for key, val in data.get('config', {}).items():
        c.execute('INSERT OR REPLACE INTO config (key, value) VALUES (?, ?)', (key, str(val)))
    # Category rules
    if 'category_rules' in data:
        c.execute('DELETE FROM category_rules')
        for rule in data['category_rules']:
            c.execute('INSERT INTO category_rules (category, base_price, min_per_team, max_per_team) VALUES (?, ?, ?, ?)',
                      (rule['category'], rule.get('base_price', 0), rule.get('min_per_team', 0), rule.get('max_per_team', 99)))
    # Teams sync if provided in wizard
    if 'teams' in data and data['teams']:
        existing = c.execute('SELECT id, name FROM teams').fetchall()
        existing_names = {t['name']: t['id'] for t in existing}
        for t in data['teams']:
            if t['name'] in existing_names:
                c.execute('UPDATE teams SET total_budget=?, color=? WHERE id=?',
                          (t['total_budget'], t.get('color', '#3b82f6'), existing_names[t['name']]))
            else:
                c.execute('INSERT INTO teams (name, total_budget, remaining_budget, color) VALUES (?, ?, ?, ?)',
                          (t['name'], t['total_budget'], t['total_budget'], t.get('color', '#3b82f6')))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/setup/restart', methods=['POST'])
def restart_setup():
    data = request.json or {}
    wipe_all = data.get('wipe_all', False)
    if wipe_all and data.get('password') != 'Wipe@123':
        return jsonify({'error': 'Incorrect password'}), 403
    # Keep a dated copy of the outgoing auction before erasing it.
    archive_excel_backup('wipe' if wipe_all else 'restart')
    conn = get_db()
    c = conn.cursor()
    c.execute('INSERT OR REPLACE INTO config (key, value) VALUES ("setup_done", "false")')
    c.execute('DELETE FROM auction_state')
    c.execute('DELETE FROM action_history')
    if wipe_all:
        c.execute('DELETE FROM players')
        c.execute('DELETE FROM teams')
        c.execute('DELETE FROM category_rules')
    else:
        c.execute('UPDATE players SET status="unsold", team_id=NULL, sold_price=NULL, sold_at=NULL')
        c.execute('UPDATE teams SET remaining_budget=total_budget')
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ─── Helper: Max Allowed Bid (reserve enough purse for remaining required slots) ───
def compute_max_bid(rem_budget, current_count, rules_list, config_dict):
    """A team must keep enough purse to fill its remaining required squad slots at
    base price. Max bid on the current player = purse − (remaining slots after this
    one) × base price. Returns (max_bid, target_squad, needed_players, reserved_spots,
    reserved_purse, common_bp)."""
    total_min_required = sum(r['min_per_team'] for r in rules_list)
    configured_min = int(config_dict.get('min_players_per_team') or config_dict.get('target_squad_size') or 0)
    target_squad = configured_min if configured_min > 0 else (total_min_required if total_min_required > 0 else 10)
    common_bp = float(config_dict.get('common_base_price') or config_dict.get('default_base_price') or 10.0)
    needed_players = max(0, target_squad - current_count)
    if needed_players <= 1:
        return round(rem_budget, 1), target_squad, needed_players, 0, 0.0, common_bp
    reserved_spots = needed_players - 1
    reserved_purse = round(reserved_spots * common_bp, 1)
    max_bid = max(0.0, round(rem_budget - reserved_purse, 1))
    # Feasibility floor: a team must always be able to buy the current player at
    # (at least) base price if it can afford it — otherwise a budget that is too
    # small for the full squad makes every player unsellable. When the reserve
    # would push the cap below base price, allow up to base price instead.
    if rem_budget >= common_bp and max_bid < common_bp:
        max_bid = common_bp
        reserved_purse = round(rem_budget - max_bid, 1)
    return max_bid, target_squad, needed_players, reserved_spots, reserved_purse, common_bp


# ─── Helper: Format Team Metrics with Max Allowed Bid & Reserved Purse ───
def format_team_metrics(team_dict, players_list, rules_list, config_dict):
    td = dict(team_dict)
    players = [dict(p) for p in players_list]
    td['players'] = players
    td['player_count'] = len(players)

    cat_counts = {}
    for p in players:
        cat = p['category'] or 'Unknown'
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    fulfillment = []
    total_min_required = 0
    for rule in rules_list:
        have = cat_counts.get(rule['category'], 0)
        min_req = rule['min_per_team']
        max_req = rule['max_per_team']
        total_min_required += min_req
        fulfillment.append({
            'category': rule['category'],
            'have': have,
            'min': min_req,
            'max': max_req,
            'met': have >= min_req,
            'exceeded': have > min_req,
            'at_max': have >= max_req
        })
    td['fulfillment'] = fulfillment

    # Max Allowed Bid — single source of truth (shared with sell enforcement)
    rem_budget = float(td['remaining_budget'])
    max_bid, target_squad, needed_players, reserved_spots, reserved_purse, common_bp = \
        compute_max_bid(rem_budget, len(players), rules_list, config_dict)

    td['target_squad_size'] = target_squad
    td['needed_players'] = needed_players
    td['common_base_price'] = common_bp
    td['max_allowed_bid'] = max_bid
    td['reserved_spots'] = reserved_spots
    td['reserved_purse'] = reserved_purse
    td['max_bid_explanation'] = f"Purse ₹{rem_budget}L − Reserved (₹{common_bp}L × {reserved_spots} spots = ₹{reserved_purse}L) = Max Bid: ₹{max_bid}L"
    return td

# ─── Teams ───
@app.route('/api/teams', methods=['GET'])
def get_teams():
    conn = get_db()
    teams = conn.execute('SELECT * FROM teams').fetchall()
    rules = conn.execute('SELECT * FROM category_rules').fetchall()
    rows = conn.execute('SELECT * FROM config').fetchall()
    config = {r['key']: r['value'] for r in rows}
    result = []
    for t in teams:
        players = conn.execute('SELECT id, name, category, sold_price, photo_url FROM players WHERE team_id = ?', (t['id'],)).fetchall()
        result.append(format_team_metrics(t, players, rules, config))
    conn.close()
    return jsonify(result)

@app.route('/api/teams', methods=['POST'])
def add_team():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    if USE_PG:
        c.execute('INSERT INTO teams (name, total_budget, remaining_budget, color) VALUES (?, ?, ?, ?) RETURNING id',
                  (data['name'], data['total_budget'], data['total_budget'], data.get('color', '#3b82f6')))
        tid = c.fetchone()['id']
    else:
        c.execute('INSERT INTO teams (name, total_budget, remaining_budget, color) VALUES (?, ?, ?, ?)',
                  (data['name'], data['total_budget'], data['total_budget'], data.get('color', '#3b82f6')))
        tid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': tid, 'success': True})

@app.route('/api/teams/edit', methods=['POST'])
def edit_team():
    data = request.json or {}
    # Accept either `id` or `team_id`; a mismatch used to return a 500.
    tid = data.get('id', data.get('team_id'))
    if tid is None:
        return jsonify({'error': 'Team id required'}), 400

    conn = get_db()
    c = conn.cursor()
    row = c.execute('SELECT * FROM teams WHERE id=?', (tid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Team not found'}), 404

    name = data.get('name', row['name'])
    if not str(name or '').strip():
        conn.close()
        return jsonify({'error': 'Name cannot be empty'}), 400
    try:
        total = float(data['total_budget']) if data.get('total_budget') is not None else row['total_budget']
    except (TypeError, ValueError):
        conn.close()
        return jsonify({'error': 'total_budget must be a number'}), 400
    color = data.get('color', row['color'] or '#3b82f6')

    # Changing the total purse shifts the remaining purse by the same delta,
    # so money already committed to signed players is preserved.
    c.execute('UPDATE teams SET name=?, remaining_budget=remaining_budget+(? - total_budget), total_budget=?, color=? WHERE id=?',
              (name, total, total, color, tid))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ─── Players ───
@app.route('/api/players', methods=['GET'])
def get_players():
    conn = get_db()
    players = conn.execute('SELECT p.*, t.name as team_name FROM players p LEFT JOIN teams t ON p.team_id = t.id').fetchall()
    conn.close()
    res = []
    for p in players:
        pd = dict(p)
        if pd.get('attributes'):
            try:
                pd['attributes'] = json.loads(pd['attributes'])
            except Exception:
                pd['attributes'] = {}
        else:
            pd['attributes'] = {}
        res.append(pd)
    return jsonify(res)

@app.route('/api/players', methods=['POST'])
def add_player():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    if USE_PG:
        c.execute('INSERT INTO players (name, category, base_price, photo_url) VALUES (?, ?, ?, ?) RETURNING id',
                  (data['name'], data.get('category', ''), data.get('base_price', 0), data.get('photo_url', '')))
        pid = c.fetchone()['id']
    else:
        c.execute('INSERT INTO players (name, category, base_price, photo_url) VALUES (?, ?, ?, ?)',
                  (data['name'], data.get('category', ''), data.get('base_price', 0), data.get('photo_url', '')))
        pid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': pid, 'success': True})

@app.route('/api/players/edit', methods=['POST'])
def edit_player():
    data = request.json or {}
    # Accept either `id` or `player_id` — the rest of the API uses `player_id`,
    # and a mismatch used to raise a KeyError and return a 500.
    pid = data.get('id', data.get('player_id'))
    if pid is None:
        return jsonify({'error': 'Player id required'}), 400

    conn = get_db()
    row = conn.execute('SELECT * FROM players WHERE id=?', (pid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Player not found'}), 404

    # Only overwrite the fields actually supplied; leave the rest untouched.
    name = data.get('name', row['name'])
    category = data.get('category', row['category'])
    try:
        base_price = float(data['base_price']) if data.get('base_price') is not None else row['base_price']
    except (TypeError, ValueError):
        conn.close()
        return jsonify({'error': 'base_price must be a number'}), 400
    if not str(name or '').strip():
        conn.close()
        return jsonify({'error': 'Name cannot be empty'}), 400

    if data.get('photo_url') is not None:
        conn.execute('UPDATE players SET name=?, category=?, base_price=?, photo_url=? WHERE id=?',
                     (name, category, base_price, data['photo_url'], pid))
    else:
        conn.execute('UPDATE players SET name=?, category=?, base_price=? WHERE id=?',
                     (name, category, base_price, pid))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/players/photo/<int:player_id>', methods=['POST'])
def upload_photo(player_id):
    if 'photo' not in request.files:
        return jsonify({'error': 'No photo'}), 400
    file = request.files['photo']
    if file.filename == '':
        return jsonify({'error': 'No file'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
    fname = f"player_{player_id}_{uuid.uuid4().hex[:8]}.{ext}"
    fpath = os.path.join(UPLOAD_FOLDER, fname)
    file.save(fpath)
    url = f"/uploads/{fname}"
    conn = get_db()
    conn.execute('UPDATE players SET photo_url=? WHERE id=?', (url, player_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'photo_url': url})

def read_raw_auction_file(file_storage_or_path):
    """
    Reads an Excel (.xlsx, .xls) or CSV (.csv) file.
    Returns (headers, dict_rows, raw_rows).
    """
    if isinstance(file_storage_or_path, str):
        filename = file_storage_or_path.lower()
        fileobj = open(file_storage_or_path, 'rb')
    else:
        filename = (file_storage_or_path.filename or '').lower()
        fileobj = file_storage_or_path

    raw_rows = []
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        import openpyxl
        wb = openpyxl.load_workbook(fileobj, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row and any(c is not None and str(c).strip() != '' for c in row):
                raw_rows.append([c if c is not None else '' for c in row])
        if isinstance(file_storage_or_path, str):
            fileobj.close()
    else:
        if isinstance(file_storage_or_path, str):
            fileobj.close()
            with open(file_storage_or_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row in reader:
                    if row and any(str(c).strip() != '' for c in row):
                        raw_rows.append(row)
        else:
            content = fileobj.read().decode('utf-8', errors='ignore')
            stream = io.StringIO(content)
            reader = csv.reader(stream)
            for row in reader:
                if row and any(str(c).strip() != '' for c in row):
                    raw_rows.append(row)

    if not raw_rows:
        return [], [], []

    # First row as headers
    headers = [str(h).strip() if h is not None and str(h).strip() != '' else f'Column {i+1}' for i, h in enumerate(raw_rows[0])]
    data_rows = raw_rows[1:]

    dict_rows = []
    for r in data_rows:
        row_dict = {}
        for i, h in enumerate(headers):
            val = r[i] if i < len(r) else ''
            row_dict[h] = val
        dict_rows.append(row_dict)

    return headers, dict_rows, data_rows


def inspect_file_data(headers, dict_rows):
    """
    Inspects column types, numeric ranges, unique values, and suggestions.
    """
    column_details = {}

    for h in headers:
        values = [r.get(h) for r in dict_rows if r.get(h) is not None and str(r.get(h)).strip() != '']
        num_count = 0
        numeric_vals = []
        for v in values:
            try:
                numeric_vals.append(float(v))
                num_count += 1
            except (ValueError, TypeError):
                pass

        is_numeric = len(values) > 0 and (num_count / len(values)) >= 0.70
        h_lower = h.lower()

        detail = {
            'name': h,
            'type': 'numeric' if is_numeric else 'text',
            'sample_values': [str(v) for v in values[:6]],
            'non_empty_count': len(values),
            'is_name_candidate': any(k in h_lower for k in ['name', 'player', 'item', 'athlete', 'title', 'lot']),
            'is_age_candidate': any(k in h_lower for k in ['age', 'yr', 'dob', 'birth']) or (is_numeric and numeric_vals and 10 <= min(numeric_vals) and max(numeric_vals) <= 100),
            'is_gender_candidate': any(k in h_lower for k in ['gender', 'sex']),
            'is_role_candidate': any(k in h_lower for k in ['role', 'cat', 'position', 'type', 'class', 'dept', 'genre', 'tier']),
            'is_price_candidate': any(k in h_lower for k in ['price', 'base', 'cost', 'reserve', 'starting', 'bid', 'value']),
            'is_photo_candidate': any(k in h_lower for k in ['photo', 'image', 'picture', 'avatar', 'img', 'url', 'link'])
        }

        if is_numeric and numeric_vals:
            detail['min'] = min(numeric_vals)
            detail['max'] = max(numeric_vals)
            detail['has_decimals'] = any(v % 1 != 0 for v in numeric_vals)
        else:
            unique_vals = []
            seen = set()
            for v in values:
                s = str(v).strip()
                if s and s not in seen:
                    seen.add(s)
                    unique_vals.append(s)
            detail['unique_values'] = unique_vals[:60]
            detail['unique_count'] = len(seen)

        column_details[h] = detail

    return column_details


def check_single_condition(row_dict, cond):
    col = cond.get('column')
    op = cond.get('operator', '==')
    target_val = cond.get('value')
    if not col or col not in row_dict:
        return False
    row_val = row_dict.get(col)
    if row_val is None or str(row_val).strip() == '':
        return False

    if op in ['==', 'equals']:
        return str(row_val).strip().lower() == str(target_val).strip().lower()
    elif op in ['!=', 'not_equals']:
        return str(row_val).strip().lower() != str(target_val).strip().lower()
    elif op == 'contains':
        return str(target_val).strip().lower() in str(row_val).strip().lower()
    elif op in ['>=', 'gte', '<=', 'lte', '>', 'gt', '<', 'lt']:
        try:
            num_r = float(row_val)
            num_t = float(target_val)
            if op in ['>=', 'gte']: return num_r >= num_t
            elif op in ['<=', 'lte']: return num_r <= num_t
            elif op in ['>', 'gt']: return num_r > num_t
            elif op in ['<', 'lt']: return num_r < num_t
        except (ValueError, TypeError):
            return False
    elif op in ['range', 'between']:
        try:
            num_r = float(row_val)
            min_v = float(cond.get('min', -999999))
            max_v = float(cond.get('max', 999999))
            return min_v <= num_r <= max_v
        except (ValueError, TypeError):
            return False
    return False


def evaluate_row_category(row_dict, rule_config, default_category='General'):
    mode = (rule_config or {}).get('mode', 'column')
    fallback = (rule_config or {}).get('fallback', default_category) or default_category

    if mode in ['multi_attribute', 'priority_rules', 'custom_rules']:
        rules = (rule_config or {}).get('rules', [])
        for r in rules:
            target_cat = (r.get('category') or r.get('name') or '').strip()
            if not target_cat:
                continue

            conditions = r.get('conditions')
            if not conditions:
                if r.get('column'):
                    conditions = [{
                        'column': r.get('column'),
                        'operator': r.get('operator', '=='),
                        'value': r.get('value'),
                        'min': r.get('min'),
                        'max': r.get('max')
                    }]
                else:
                    continue

            if all(check_single_condition(row_dict, c) for c in conditions):
                return target_cat

        return fallback

    elif mode == 'column':
        col = (rule_config or {}).get('column')
        if not col or col not in row_dict:
            return fallback
        val = str(row_dict.get(col) or '').strip()
        return val if val else fallback

    elif mode == 'numeric_brackets':
        col = (rule_config or {}).get('column')
        if not col or col not in row_dict:
            return fallback
        try:
            val = float(row_dict.get(col))
        except (ValueError, TypeError):
            return fallback

        brackets = (rule_config or {}).get('brackets', [])
        for b in brackets:
            b_name = b.get('name', '').strip()
            if not b_name:
                continue
            try:
                min_v = float(b.get('min', -999999))
            except (ValueError, TypeError):
                min_v = -999999
            try:
                max_v = float(b.get('max', 999999))
            except (ValueError, TypeError):
                max_v = 999999
            if min_v <= val <= max_v:
                return b_name
        return fallback

    elif mode == 'manual':
        return (rule_config or {}).get('default_category', default_category) or default_category

    return fallback


def parse_auction_file(file_storage, default_base_price=10.0):
    """Legacy parser preserved for simple imports"""
    headers, dict_rows, _ = read_raw_auction_file(file_storage)
    if not dict_rows:
        return []

    col_details = inspect_file_data(headers, dict_rows)
    name_col = next((h for h in headers if col_details[h]['is_name_candidate']), headers[0])
    cat_col = next((h for h in headers if col_details[h]['is_role_candidate'] and h != name_col), None)
    bp_col = next((h for h in headers if col_details[h]['is_price_candidate']), None)
    photo_col = next((h for h in headers if col_details[h]['is_photo_candidate']), None)

    items = []
    for r in dict_rows:
        name = str(r.get(name_col) or '').strip()
        if not name:
            continue
        cat = str(r.get(cat_col) or 'General').strip() if cat_col else 'General'
        bp = default_base_price
        if bp_col and r.get(bp_col) is not None:
            try:
                bp = float(r.get(bp_col))
            except (ValueError, TypeError):
                bp = default_base_price
        photo = str(r.get(photo_col) or '').strip() if photo_col else ''
        attrs = {k: v for k, v in r.items() if v is not None and str(v).strip() != '' and k not in [name_col, photo_col]}
        items.append({'name': name, 'category': cat, 'base_price': bp, 'photo_url': photo, 'attributes': attrs})

    return items


# ─── New Dynamic File Ingestion & Category Rule Endpoints ───
@app.route('/api/file/inspect', methods=['POST'])
def inspect_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty filename'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'xlsx'
    file_id = uuid.uuid4().hex
    saved_filename = f"temp_{file_id}.{ext}"
    saved_path = os.path.join(UPLOAD_FOLDER, saved_filename)
    file.save(saved_path)

    try:
        headers, dict_rows, data_rows = read_raw_auction_file(saved_path)
    except Exception as e:
        return jsonify({'error': f'Failed to read file: {str(e)}'}), 400

    if not dict_rows:
        return jsonify({'error': 'File contains no data rows'}), 400

    column_details = inspect_file_data(headers, dict_rows)

    # Suggest best default column mappings
    suggested_name_col = next((h for h in headers if column_details[h]['is_name_candidate']), headers[0])
    suggested_cat_col = next((h for h in headers if (column_details[h]['is_role_candidate'] or column_details[h]['is_gender_candidate']) and h != suggested_name_col), None)
    suggested_age_col = next((h for h in headers if column_details[h]['is_age_candidate']), None)
    suggested_price_col = next((h for h in headers if column_details[h]['is_price_candidate']), None)
    suggested_photo_col = next((h for h in headers if column_details[h]['is_photo_candidate']), None)

    return jsonify({
        'success': True,
        'file_id': file_id,
        'filename': file.filename,
        'total_rows': len(dict_rows),
        'columns': headers,
        'column_details': column_details,
        'suggestions': {
            'name_column': suggested_name_col,
            'category_column': suggested_cat_col,
            'age_column': suggested_age_col,
            'price_column': suggested_price_col,
            'photo_column': suggested_photo_col
        },
        'preview_rows': dict_rows[:10]
    })


@app.route('/api/file/preview_categorization', methods=['POST'])
def preview_categorization():
    data = request.json or {}
    file_id = data.get('file_id')
    preset_id = data.get('preset_id')

    if file_id:
        matching_files = [f for f in os.listdir(UPLOAD_FOLDER) if f.startswith(f"temp_{file_id}.")]
        if not matching_files:
            return jsonify({'error': 'Uploaded file session expired. Please re-upload.'}), 404
        file_path = os.path.join(UPLOAD_FOLDER, matching_files[0])
    elif preset_id:
        preset_files = {
            'football': 'test_football.xlsx',
            'basketball': 'test_basketball.xlsx',
            'esports': 'test_esports.xlsx',
            'art': 'test_art_antiques.xlsx',
            'cricket': 'test_players.csv',
            'multi_age': 'test_multi_sport_and_age.xlsx'
        }
        fname = preset_files.get(preset_id, 'test_multi_sport_and_age.xlsx')
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)
    else:
        return jsonify({'error': 'No file_id or preset_id provided'}), 400

    try:
        headers, dict_rows, _ = read_raw_auction_file(file_path)
    except Exception as e:
        return jsonify({'error': f'Failed to parse file: {str(e)}'}), 400

    name_col = data.get('name_column') or headers[0]
    photo_col = data.get('photo_column')
    base_price_col = data.get('base_price_column')
    default_bp = float(data.get('default_base_price', 10))
    rule_config = data.get('rule_config', {'mode': 'column', 'column': headers[1] if len(headers) > 1 else headers[0]})

    cat_counts = {}
    categorized_samples = {}

    for r in dict_rows:
        name = str(r.get(name_col) or '').strip()
        if not name:
            continue
        cat = evaluate_row_category(r, rule_config, default_category='General')
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

        bp = default_bp
        if base_price_col and r.get(base_price_col) is not None:
            try:
                bp = float(r.get(base_price_col))
            except (ValueError, TypeError):
                bp = default_bp

        photo = str(r.get(photo_col) or '').strip() if photo_col else ''
        row_attrs = {k: v for k, v in r.items() if v is not None and str(v).strip() != '' and k not in [name_col, photo_col]}

        if cat not in categorized_samples:
            categorized_samples[cat] = []
        if len(categorized_samples[cat]) < 5:
            categorized_samples[cat].append({
                'name': name,
                'category': cat,
                'base_price': bp,
                'photo_url': photo,
                'attributes': row_attrs,
                'raw': r
            })

    summary = []
    for cat_name, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
        summary.append({
            'category': cat_name,
            'count': count,
            'samples': categorized_samples.get(cat_name, [])
        })

    return jsonify({
        'success': True,
        'total_items': sum(cat_counts.values()),
        'categories_count': len(cat_counts),
        'categories': summary
    })


@app.route('/api/file/apply_and_launch', methods=['POST'])
def apply_and_launch():
    data = request.json or {}
    file_id = data.get('file_id')
    preset_id = data.get('preset_id')

    if file_id:
        matching_files = [f for f in os.listdir(UPLOAD_FOLDER) if f.startswith(f"temp_{file_id}.")]
        if not matching_files:
            return jsonify({'error': 'Uploaded file session expired. Please re-upload.'}), 404
        file_path = os.path.join(UPLOAD_FOLDER, matching_files[0])
    elif preset_id:
        preset_files = {
            'football': 'test_football.xlsx',
            'basketball': 'test_basketball.xlsx',
            'esports': 'test_esports.xlsx',
            'art': 'test_art_antiques.xlsx',
            'cricket': 'test_players.csv',
            'multi_age': 'test_multi_sport_and_age.xlsx'
        }
        fname = preset_files.get(preset_id, 'test_multi_sport_and_age.xlsx')
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)
    else:
        return jsonify({'error': 'No file_id or preset_id provided'}), 400

    try:
        headers, dict_rows, _ = read_raw_auction_file(file_path)
    except Exception as e:
        return jsonify({'error': f'Failed to parse file: {str(e)}'}), 400

    name_col = data.get('name_column') or headers[0]
    photo_col = data.get('photo_column')
    base_price_col = data.get('base_price_column')
    default_bp = float(data.get('default_base_price', 10))
    rule_config = data.get('rule_config', {'mode': 'column', 'column': headers[1] if len(headers) > 1 else headers[0]})
    user_cat_rules = {r['category']: r for r in data.get('category_rules', [])}

    conn = get_db()
    c = conn.cursor()

    # Clear previous state
    c.execute('DELETE FROM players')
    c.execute('DELETE FROM category_rules')
    c.execute('DELETE FROM teams')
    c.execute('DELETE FROM auction_state')
    c.execute('DELETE FROM action_history')

    # Save Config
    cfg_data = data.get('config', {})
    event_name = cfg_data.get('event_name', 'Premier Auction 2026')
    common_bp = float(cfg_data.get('common_base_price', default_bp))
    min_players = int(cfg_data.get('min_players_per_team', 10))
    bid_inc = float(cfg_data.get('bid_increment', 2.5))

    c.execute('INSERT OR REPLACE INTO config (key, value) VALUES ("event_name", ?)', (event_name,))
    c.execute('INSERT OR REPLACE INTO config (key, value) VALUES ("common_base_price", ?)', (str(common_bp),))
    c.execute('INSERT OR REPLACE INTO config (key, value) VALUES ("min_players_per_team", ?)', (str(min_players),))
    c.execute('INSERT OR REPLACE INTO config (key, value) VALUES ("bid_increment", ?)', (str(bid_inc),))
    c.execute('INSERT OR REPLACE INTO config (key, value) VALUES ("setup_done", "true")')

    # Insert Categorized Players
    cat_counts = {}
    player_count = 0

    for r in dict_rows:
        name = str(r.get(name_col) or '').strip()
        if not name:
            continue
        cat = evaluate_row_category(r, rule_config, default_category='General')
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

        bp = common_bp
        if base_price_col and r.get(base_price_col) is not None:
            try:
                bp = float(r.get(base_price_col))
            except (ValueError, TypeError):
                bp = common_bp

        photo = str(r.get(photo_col) or '').strip() if photo_col else ''
        row_attrs = {k: v for k, v in r.items() if v is not None and str(v).strip() != '' and k not in [name_col, photo_col]}

        c.execute('INSERT INTO players (name, category, base_price, photo_url, attributes) VALUES (?, ?, ?, ?, ?)',
                  (name, cat, bp, photo, json.dumps(row_attrs)))
        player_count += 1

    # Insert Category Rules with quotas
    saved_rules = []
    for cat_name, cnt in cat_counts.items():
        rule_info = user_cat_rules.get(cat_name, {})
        bp = float(rule_info.get('base_price', common_bp))
        min_req = int(rule_info.get('min_per_team', 0))
        max_req = int(rule_info.get('max_per_team', 99))
        c.execute('INSERT INTO category_rules (category, base_price, min_per_team, max_per_team) VALUES (?, ?, ?, ?)',
                  (cat_name, bp, min_req, max_req))
        saved_rules.append({
            'category': cat_name,
            'base_price': bp,
            'min_per_team': min_req,
            'max_per_team': max_req,
            'count': cnt
        })

    # Insert Custom Teams
    teams_data = data.get('teams', [])
    for t in teams_data:
        t_name = t.get('name', '').strip()
        if not t_name:
            continue
        budget = float(t.get('total_budget', 1000))
        color = t.get('color', '#3b82f6')
        c.execute('INSERT INTO teams (name, total_budget, remaining_budget, color) VALUES (?, ?, ?, ?)',
                  (t_name, budget, budget, color))

    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'player_count': player_count,
        'categories': saved_rules,
        'teams_count': len(teams_data),
        'event_name': event_name
    })


@app.route('/api/players/import', methods=['POST'])
def import_players():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty filename'}), 400

    conn = get_db()
    c = conn.cursor()
    cfg_bp_row = c.execute('SELECT value FROM config WHERE key="common_base_price"').fetchone()
    default_bp = float(cfg_bp_row['value']) if cfg_bp_row else 10.0

    try:
        items = parse_auction_file(file, default_base_price=default_bp)
    except Exception as e:
        conn.close()
        return jsonify({'error': f'Failed to parse file: {str(e)}'}), 400

    if not items:
        conn.close()
        return jsonify({'error': 'No valid items found in file'}), 400

    cat_counts = {}
    for it in items:
        cat = it['category'] or 'General'
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    count = 0
    for it in items:
        c.execute('INSERT INTO players (name, category, base_price, photo_url, attributes) VALUES (?, ?, ?, ?, ?)',
                  (it['name'], it['category'], it['base_price'], it['photo_url'], json.dumps(it.get('attributes') or {})))
        count += 1

    existing_rules = {r['category']: dict(r) for r in c.execute('SELECT * FROM category_rules').fetchall()}
    detected_rules = []

    for cat_name, c_cnt in cat_counts.items():
        if cat_name in existing_rules:
            rule = existing_rules[cat_name]
            rule['count'] = c_cnt
            detected_rules.append(rule)
        else:
            c.execute('INSERT INTO category_rules (category, base_price, min_per_team, max_per_team) VALUES (?, ?, ?, ?)',
                      (cat_name, default_bp, 0, 99))
            detected_rules.append({
                'category': cat_name,
                'base_price': default_bp,
                'min_per_team': 0,
                'max_per_team': 99,
                'count': c_cnt
            })

    all_rules = [dict(r) for r in c.execute('SELECT * FROM category_rules').fetchall()]

    conn.commit()
    conn.close()
    return jsonify({
        'success': True,
        'count': count,
        'detected_categories': detected_rules,
        'all_category_rules': all_rules
    })


@app.route('/api/players/clear_pool', methods=['POST'])
def clear_player_pool():
    conn = get_db()
    c = conn.cursor()
    c.execute('DELETE FROM players')
    c.execute('DELETE FROM category_rules')
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/presets/list', methods=['GET'])
def list_presets():
    presets = [
        {'id': 'multi_age', 'name': '⭐ Multi-Sport with Age & Gender (38 Stars)', 'file': 'test_multi_sport_and_age.xlsx', 'event_name': 'Superstars All-Star Auction 2026', 'columns': ['Name', 'Age', 'Gender', 'Role', 'Base Price', 'Experience (Yrs)', 'City', 'Photo URL']},
        {'id': 'football', 'name': '⚽ Football (Soccer)', 'file': 'test_football.xlsx', 'event_name': 'Super Football League 2026', 'categories': ['Forward', 'Midfielder', 'Defender', 'Goalkeeper', 'Winger']},
        {'id': 'basketball', 'name': '🏀 Basketball', 'file': 'test_basketball.xlsx', 'event_name': 'Pro Basketball Draft 2026', 'categories': ['Point Guard', 'Shooting Guard', 'Small Forward', 'Power Forward', 'Center']},
        {'id': 'esports', 'name': '🎮 Esports / Gaming', 'file': 'test_esports.xlsx', 'event_name': 'Champions Esports Auction 2026', 'categories': ['Duelist', 'Initiator', 'Controller', 'Sentinel', 'IGL']},
        {'id': 'art', 'name': '🎨 Fine Art & Antiques', 'file': 'test_art_antiques.xlsx', 'event_name': 'Grand Heritage Art & Antiques', 'categories': ['Oil Painting', 'Sculpture', 'Vintage Watch', 'Rare Coin', 'Antique Furniture']},
        {'id': 'cricket', 'name': '🏏 Cricket 100 Players', 'file': 'test_players.csv', 'event_name': 'Premier Cricket League 2026', 'categories': ['Batsman', 'Bowler', 'All-Rounder', 'Wicket-Keeper', 'Female', '50+', 'Under-19']}
    ]
    return jsonify({'presets': presets})

@app.route('/api/presets/load', methods=['POST'])
def load_preset():
    preset_id = (request.json or {}).get('preset_id', 'football')
    preset_files = {
        'football': ('test_football.xlsx', 'Super Football League 2026'),
        'basketball': ('test_basketball.xlsx', 'Pro Basketball Draft 2026'),
        'esports': ('test_esports.xlsx', 'Champions Esports Auction 2026'),
        'art': ('test_art_antiques.xlsx', 'Grand Heritage Art & Antiques'),
        'cricket': ('test_players.csv', 'Premier Cricket League 2026')
    }

    if preset_id not in preset_files:
        return jsonify({'error': 'Preset not found'}), 404

    fname, event_name = preset_files[preset_id]
    fpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)
    if not os.path.exists(fpath):
        return jsonify({'error': f'File {fname} not found'}), 404

    conn = get_db()
    c = conn.cursor()

    # Clear old pool and categories for a clean preset load
    c.execute('DELETE FROM players')
    c.execute('DELETE FROM category_rules')
    c.execute('UPDATE config SET value=? WHERE key="event_name"', (event_name,))

    cfg_bp_row = c.execute('SELECT value FROM config WHERE key="common_base_price"').fetchone()
    default_bp = float(cfg_bp_row['value']) if cfg_bp_row else 10.0

    count = 0
    cat_counts = {}

    if fname.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(fpath, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        for r in rows[1:]:
            if not r or not r[0]: continue
            name = str(r[0]).strip()
            cat = str(r[1]).strip() if len(r) > 1 and r[1] else 'General'
            bp = float(r[2]) if len(r) > 2 and r[2] is not None else default_bp
            photo = str(r[3]).strip() if len(r) > 3 and r[3] else ''
            c.execute('INSERT INTO players (name, category, base_price, photo_url) VALUES (?, ?, ?, ?)', (name, cat, bp, photo))
            count += 1
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
    elif fname.endswith('.csv'):
        with open(fpath, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader, None)
            for r in reader:
                if not r or not r[0].strip(): continue
                name = r[0].strip()
                cat = r[1].strip() if len(r) > 1 else 'General'
                bp = float(r[2].strip()) if len(r) > 2 and r[2].strip() else default_bp
                photo = r[3].strip() if len(r) > 3 else ''
                c.execute('INSERT INTO players (name, category, base_price, photo_url) VALUES (?, ?, ?, ?)', (name, cat, bp, photo))
                count += 1
                cat_counts[cat] = cat_counts.get(cat, 0) + 1

    # Auto-populate category_rules
    detected_rules = []
    for cat_name, c_cnt in cat_counts.items():
        c.execute('INSERT INTO category_rules (category, base_price, min_per_team, max_per_team) VALUES (?, ?, ?, ?)',
                  (cat_name, default_bp, 0, 99))
        detected_rules.append({
            'category': cat_name,
            'base_price': default_bp,
            'min_per_team': 0,
            'max_per_team': 99,
            'count': c_cnt
        })

    all_rules = [dict(r) for r in c.execute('SELECT * FROM category_rules').fetchall()]

    conn.commit()
    conn.close()
    return jsonify({
        'success': True,
        'count': count,
        'event_name': event_name,
        'detected_categories': detected_rules,
        'all_category_rules': all_rules
    })

@app.route('/api/players/load_test_data', methods=['POST'])
def load_test_data():
    return load_preset()

# ─── Auction Actions ───
@app.route('/api/auction/state', methods=['GET'])
def get_auction_state():
    conn = get_db()
    rows = conn.execute('SELECT * FROM auction_state').fetchall()
    state = {r['key']: r['value'] for r in rows}
    if state.get('current_player'):
        p_row = conn.execute('SELECT * FROM players WHERE name = ?', (state['current_player'],)).fetchone()
        if p_row and p_row['attributes']:
            state['attributes'] = p_row['attributes']
    conn.close()
    return jsonify(state)

@app.route('/api/auction/state', methods=['POST'])
def set_auction_state():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    for k, v in data.items():
        c.execute('INSERT OR REPLACE INTO auction_state (key, value) VALUES (?, ?)', (k, str(v)))
    # Clear sold/passed overlay and any previous leading bidder when a new
    # player goes on the block (unless this same request sets a bidder).
    if data.get('current_player'):
        keys = ['auction_status', 'last_sold_player', 'last_sold_price', 'last_sold_team_name',
                'last_sold_team_color', 'last_sold_photo', 'last_passed_player', 'last_passed_photo']
        if 'bidder_team_id' not in data:
            keys += ['bidder_team_id', 'bidder_team_name', 'bidder_team_color']
        c.execute("DELETE FROM auction_state WHERE key IN (%s)" % ','.join(['?'] * len(keys)), keys)
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/sell_player', methods=['POST'])
def sell_player():
    data = request.json or {}
    pid, tid = data.get('player_id'), data.get('team_id')
    if pid is None or tid is None:
        return jsonify({'error': 'player_id and team_id are required'}), 400
    try:
        price = float(data.get('sold_price'))
    except (TypeError, ValueError):
        return jsonify({'error': 'sold_price must be a number'}), 400
    if price < 0:
        return jsonify({'error': 'sold_price cannot be negative'}), 400

    conn = get_db()
    c = conn.cursor()
    player = c.execute('SELECT * FROM players WHERE id=?', (pid,)).fetchone()
    if not player:
        conn.close()
        return jsonify({'error': 'Player not found'}), 404
    team = c.execute('SELECT * FROM teams WHERE id=?', (tid,)).fetchone()
    if not team:
        conn.close()
        return jsonify({'error': 'Team not found'}), 404
    # Selling an already-sold player would double-charge and corrupt undo history
    if player['status'] == 'sold':
        conn.close()
        return jsonify({'error': '%s is already sold. Undo the previous sale first.' % player['name']}), 409

    # ── Enforce Max Allowed Bid (reserve purse for remaining required slots) ──
    # This also guarantees the team's purse can never go negative.
    rules = c.execute('SELECT * FROM category_rules').fetchall()
    cfg_rows = c.execute('SELECT * FROM config').fetchall()
    cfg = {r['key']: r['value'] for r in cfg_rows}
    current_count = c.execute("SELECT COUNT(*) as n FROM players WHERE team_id=? AND status='sold'", (tid,)).fetchone()['n']
    rem_budget = float(team['remaining_budget'])
    max_bid, _tsq, _need, _rspots, reserved_purse, common_bp = \
        compute_max_bid(rem_budget, current_count, [dict(r) for r in rules], cfg)
    if price > max_bid + 1e-6:
        conn.close()
        if reserved_purse > 0:
            msg = ('Bid ₹%gL exceeds %s\'s max allowed bid of ₹%gL. '
                   'They must keep ₹%gL reserved to fill remaining squad slots at ₹%gL base each.'
                   % (price, team['name'], max_bid, reserved_purse, common_bp))
        else:
            msg = ('Bid ₹%gL exceeds %s\'s remaining purse of ₹%gL.'
                   % (price, team['name'], max_bid))
        return jsonify({'error': msg}), 400

    # Record in action_history for multi-level undo
    c.execute('''INSERT INTO action_history 
        (action_type, player_id, player_name, old_team_id, new_team_id, old_sold_price, new_sold_price, old_status, new_status, base_price, category, photo_url)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        ('sell', pid, player['name'], player['team_id'], tid, player['sold_price'], price, player['status'], 'sold', player['base_price'], player['category'], player['photo_url']))

    c.execute('UPDATE players SET status="sold", team_id=?, sold_price=?, sold_at=CURRENT_TIMESTAMP WHERE id=?', (tid, price, pid))
    c.execute('UPDATE teams SET remaining_budget=remaining_budget-? WHERE id=?', (price, tid))
    # Clear player-specific auction state but preserve stage settings (sport, template mode)
    c.execute("DELETE FROM auction_state WHERE key NOT IN ('auction_sport','auction_template_mode','auction_template')")
    # Write SOLD state so Cinematic Stage can show the sold overlay
    c.execute("INSERT OR REPLACE INTO auction_state (key, value) VALUES ('auction_status', 'sold')")
    c.execute("INSERT OR REPLACE INTO auction_state (key, value) VALUES ('last_sold_player', ?)", (player['name'],))
    c.execute("INSERT OR REPLACE INTO auction_state (key, value) VALUES ('last_sold_price', ?)", (str(price),))
    c.execute("INSERT OR REPLACE INTO auction_state (key, value) VALUES ('last_sold_team_name', ?)", (team['name'] if team else '',))
    c.execute("INSERT OR REPLACE INTO auction_state (key, value) VALUES ('last_sold_team_color', ?)", (team['color'] if team else '#3b82f6',))
    c.execute("INSERT OR REPLACE INTO auction_state (key, value) VALUES ('last_sold_photo', ?)", (player['photo_url'] or '',))
    conn.commit()
    conn.close()
    excel_backup_async()
    # Push to Google Sheets in background (non-blocking)
    sheets_push_async({
        'player_name': player['name'],
        'category':    player['category'] or '',
        'base_price':  player['base_price'],
        'sold_price':  price,
        'team_name':   team['name'] if team else '',
        'sold_at':     str(__import__('datetime').datetime.now()),
    })
    return jsonify({'success': True})

@app.route('/api/undo', methods=['POST'])
def undo_last_sale():
    conn = get_db()
    c = conn.cursor()

    # Check action_history stack first
    last_action = c.execute('SELECT * FROM action_history ORDER BY id DESC LIMIT 1').fetchone()

    if last_action:
        act = dict(last_action)
        pid = act['player_id']

        if act['action_type'] in ('sell',):
            # Refund buying team
            if act['new_team_id'] and act['new_sold_price']:
                c.execute('UPDATE teams SET remaining_budget=remaining_budget+? WHERE id=?', (act['new_sold_price'], act['new_team_id']))
            # Restore player old status/team/price
            c.execute('UPDATE players SET status=?, team_id=?, sold_price=?, sold_at=NULL WHERE id=?',
                      (act['old_status'] or 'unsold', act['old_team_id'], act['old_sold_price'], pid))

        elif act['action_type'] == 'pass':
            # Restore player to unsold — no money involved
            c.execute('UPDATE players SET status="unsold", team_id=NULL, sold_price=NULL WHERE id=?', (pid,))

        elif act['action_type'] == 'edit_sale':
            # Revert team balances
            if act['new_team_id'] and act['new_sold_price']:
                c.execute('UPDATE teams SET remaining_budget=remaining_budget+? WHERE id=?', (act['new_sold_price'], act['new_team_id']))
            if act['old_team_id'] and act['old_sold_price']:
                c.execute('UPDATE teams SET remaining_budget=remaining_budget-? WHERE id=?', (act['old_sold_price'], act['old_team_id']))
            # Restore player
            c.execute('UPDATE players SET status=?, team_id=?, sold_price=? WHERE id=?',
                      (act['old_status'], act['old_team_id'], act['old_sold_price'], pid))

        # Remove this action from history stack
        c.execute('DELETE FROM action_history WHERE id=?', (act['id'],))

        # Restore player to live auction_state so they immediately appear under the hammer!
        c.execute("DELETE FROM auction_state WHERE key NOT IN ('auction_sport','auction_template_mode','auction_template')")
        c.execute('INSERT OR REPLACE INTO auction_state (key, value) VALUES ("current_player", ?)', (act['player_name'] or '',))
        restored_bid = act['new_sold_price'] if act['new_sold_price'] is not None else act['base_price'] or 0
        c.execute('INSERT INTO auction_state (key, value) VALUES ("current_bid", ?)', (str(restored_bid),))
        c.execute('INSERT INTO auction_state (key, value) VALUES ("category", ?)', (act['category'] or '',))
        c.execute('INSERT INTO auction_state (key, value) VALUES ("base_price", ?)', (str(act['base_price'] or 0),))
        c.execute('INSERT INTO auction_state (key, value) VALUES ("photo_url", ?)', (act['photo_url'] or '',))

        player_row = c.execute('SELECT * FROM players WHERE id=?', (pid,)).fetchone()
        player_dict = dict(player_row) if player_row else {'id': pid, 'name': act['player_name'], 'category': act['category'], 'base_price': act['base_price'], 'photo_url': act['photo_url']}

        rem_history = c.execute('SELECT COUNT(*) as c FROM action_history').fetchone()['c']
        conn.commit()
        conn.close()
        excel_backup_async()
        return jsonify({
            'success': True,
            'player': player_dict,
            'restored_bid': restored_bid,
            'player_name': act['player_name'],
            'remaining_undos': rem_history
        })
    else:
        # Fallback to last sold player in players table
        last = c.execute('SELECT * FROM players WHERE status="sold" AND sold_at IS NOT NULL ORDER BY sold_at DESC LIMIT 1').fetchone()
        if not last:
            conn.close()
            return jsonify({'success': False, 'error': 'Nothing to undo'})

        c.execute('UPDATE teams SET remaining_budget=remaining_budget+? WHERE id=?', (last['sold_price'], last['team_id']))
        c.execute('UPDATE players SET status="unsold", team_id=NULL, sold_price=NULL, sold_at=NULL WHERE id=?', (last['id'],))

        # Restore to auction state
        c.execute("DELETE FROM auction_state WHERE key NOT IN ('auction_sport','auction_template_mode','auction_template')")
        c.execute('INSERT OR REPLACE INTO auction_state (key, value) VALUES ("current_player", ?)', (last['name'],))
        restored_bid = last['sold_price'] or last['base_price'] or 0
        c.execute('INSERT INTO auction_state (key, value) VALUES ("current_bid", ?)', (str(restored_bid),))
        c.execute('INSERT INTO auction_state (key, value) VALUES ("category", ?)', (last['category'] or '',))
        c.execute('INSERT INTO auction_state (key, value) VALUES ("base_price", ?)', (str(last['base_price'] or 0),))
        c.execute('INSERT INTO auction_state (key, value) VALUES ("photo_url", ?)', (last['photo_url'] or '',))

        player_dict = dict(last)
        player_dict['status'] = 'unsold'
        player_dict['team_id'] = None
        player_dict['sold_price'] = None

        conn.commit()
        conn.close()
        excel_backup_async()
        return jsonify({
            'success': True,
            'player': player_dict,
            'restored_bid': restored_bid,
            'player_name': last['name'],
            'remaining_undos': 0
        })

@app.route('/api/players/edit_sale', methods=['POST'])
def edit_player_sale():
    data = request.json
    pid = data.get('player_id')
    name = data.get('name')
    category = data.get('category')
    base_price = float(data.get('base_price', 0)) if data.get('base_price') is not None else 0
    photo_url = data.get('photo_url')
    new_status = data.get('status', 'unsold') # 'sold' or 'unsold'
    new_team_id = int(data.get('team_id')) if data.get('team_id') else None
    new_sold_price = float(data.get('sold_price', 0)) if data.get('sold_price') is not None else None

    conn = get_db()
    c = conn.cursor()
    old_p = c.execute('SELECT * FROM players WHERE id=?', (pid,)).fetchone()
    if not old_p:
        conn.close()
        return jsonify({'error': 'Player not found'}), 404

    old_status = old_p['status']
    old_team_id = old_p['team_id']
    old_sold_price = old_p['sold_price'] or 0

    # Record in history for undo
    c.execute('''INSERT INTO action_history 
        (action_type, player_id, player_name, old_team_id, new_team_id, old_sold_price, new_sold_price, old_status, new_status, base_price, category, photo_url)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        ('edit_sale', pid, name or old_p['name'], old_team_id, new_team_id, old_sold_price, new_sold_price, old_status, new_status, base_price, category, photo_url or old_p['photo_url']))

    # Rebalance team budgets
    # 1. If was sold previously, refund old team
    if old_status == 'sold' and old_team_id and old_sold_price > 0:
        c.execute('UPDATE teams SET remaining_budget = remaining_budget + ? WHERE id = ?', (old_sold_price, old_team_id))

    # 2. If is sold now, deduct new team
    if new_status == 'sold' and new_team_id and new_sold_price and new_sold_price > 0:
        c.execute('UPDATE teams SET remaining_budget = remaining_budget - ? WHERE id = ?', (new_sold_price, new_team_id))

    # 3. Update player record
    c.execute('''UPDATE players SET 
        name = ?, 
        category = ?, 
        base_price = ?, 
        photo_url = COALESCE(?, photo_url),
        status = ?, 
        team_id = ?, 
        sold_price = ?,
        sold_at = CASE WHEN ? = 'sold' THEN COALESCE(sold_at, CURRENT_TIMESTAMP) ELSE NULL END
        WHERE id = ?''',
        (name or old_p['name'], category or old_p['category'], base_price, photo_url, new_status, new_team_id, new_sold_price if new_status == 'sold' else None, new_status, pid))

    conn.commit()
    conn.close()
    excel_backup_async()
    return jsonify({'success': True})

@app.route('/api/reset', methods=['POST'])
def reset_auction():
    # Archive first: clearing every sale would otherwise overwrite the backup
    # of a finished auction with an empty snapshot.
    archive_excel_backup('reset')
    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE players SET status="unsold", team_id=NULL, sold_price=NULL, sold_at=NULL')
    c.execute('UPDATE teams SET remaining_budget=total_budget')
    c.execute("DELETE FROM auction_state WHERE key NOT IN ('auction_sport','auction_template_mode','auction_template')")
    c.execute('DELETE FROM action_history')
    conn.commit()
    conn.close()
    excel_backup_async()
    return jsonify({'success': True})

# ─── Google Sheets sync ───
@app.route('/api/sheets/sync_all', methods=['POST'])
def sheets_sync_all():
    """Push every sold player to Google Sheets (full re-sync)."""
    conn = get_db()
    rows = conn.execute('''
        SELECT p.name, p.category, p.base_price, p.sold_price, t.name as team_name, p.sold_at
        FROM players p JOIN teams t ON p.team_id = t.id
        WHERE p.status = 'sold' ORDER BY p.sold_at
    ''').fetchall()
    conn.close()
    count = 0
    for r in rows:
        rd = dict(r)
        sheets_push_async({
            'player_name': rd['name'],
            'category':    rd['category'] or '',
            'base_price':  rd['base_price'],
            'sold_price':  rd['sold_price'],
            'team_name':   rd['team_name'] or '',
            'sold_at':     str(rd['sold_at'] or ''),
        })
        count += 1
    return jsonify({'success': True, 'synced': count})

# ─── DB migration (SQLite → PostgreSQL) ───
@app.route('/api/db/migrate_to_pg', methods=['POST'])
def migrate_to_pg():
    """Copy all data from local auction.db into the configured PostgreSQL database."""
    if not USE_PG:
        return jsonify({'error': 'DATABASE_URL not configured'}), 400
    try:
        import psycopg2
        src = sqlite3.connect(DB_FILE, timeout=20)
        src.row_factory = sqlite3.Row
        dst = psycopg2.connect(DATABASE_URL)
        dc = dst.cursor()

        tables = ['config', 'category_rules', 'teams', 'players', 'auction_state', 'action_history']
        counts = {}
        for tbl in tables:
            rows = src.execute(f'SELECT * FROM {tbl}').fetchall()
            if not rows:
                counts[tbl] = 0
                continue
            cols = rows[0].keys()
            col_str = ', '.join(cols)
            ph = ', '.join(['%s'] * len(cols))
            conflict = ''
            if tbl in ('config', 'auction_state'):
                conflict = ' ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value'
            dc.executemany(
                f'INSERT INTO {tbl} ({col_str}) VALUES ({ph}){conflict}',
                [tuple(r) for r in rows]
            )
            counts[tbl] = len(rows)

        # Reset sequences so auto-increment stays correct
        for tbl, col in [('teams','id'), ('players','id'), ('category_rules','id'), ('action_history','id')]:
            dc.execute(f"SELECT setval(pg_get_serial_sequence('{tbl}','{col}'), COALESCE((SELECT MAX({col}) FROM {tbl}), 1))")

        dst.commit()
        dst.close()
        src.close()
        return jsonify({'success': True, 'migrated': counts})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ─── Stats ───
@app.route('/api/stats', methods=['GET'])
def get_stats():
    conn = get_db()
    total = conn.execute('SELECT COUNT(*) as c FROM players').fetchone()['c']
    sold = conn.execute('SELECT COUNT(*) as c FROM players WHERE status="sold"').fetchone()['c']
    unsold = conn.execute("SELECT COUNT(*) as c FROM players WHERE status IN ('unsold','passed')").fetchone()['c']
    spent = conn.execute('SELECT COALESCE(SUM(sold_price),0) as s FROM players WHERE status="sold"').fetchone()['s']
    
    # Category spending analytics
    cats = conn.execute('''
        SELECT category, 
               COUNT(*) as total, 
               SUM(CASE WHEN status="sold" THEN 1 ELSE 0 END) as sold_count,
               COALESCE(SUM(CASE WHEN status="sold" THEN sold_price ELSE 0 END), 0) as total_spent,
               COALESCE(AVG(CASE WHEN status="sold" THEN sold_price ELSE NULL END), 0) as avg_price
        FROM players 
        WHERE category IS NOT NULL AND category != ''
        GROUP BY category
    ''').fetchall()
    
    cat_analytics = []
    for c in cats:
        cd = dict(c)
        cd['percent_of_total'] = round((cd['total_spent'] / spent * 100), 1) if spent > 0 else 0
        cd['avg_price'] = round(cd['avg_price'], 1)
        cat_analytics.append(cd)
        
    conn.close()
    return jsonify({
        'total_players': total,
        'sold': sold,
        'unsold': unsold,
        'total_spent': spent,
        # `live_data.stats` reports the same figures as `total`/`spent`.
        # Both spellings are served here so either convention works.
        'total': total,
        'spent': spent,
        'categories': cat_analytics
    })

# ─── Logo / Banner Uploads ───
@app.route('/api/config/banner', methods=['POST'])
def upload_banner():
    if 'banner' not in request.files:
        return jsonify({'error': 'No file'}), 400
    file = request.files['banner']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'png'
    fname = f"event_banner_{uuid.uuid4().hex[:8]}.{ext}"
    fpath = os.path.join(UPLOAD_FOLDER, fname)
    file.save(fpath)
    url = f"/uploads/{fname}"
    conn = get_db()
    conn.execute('INSERT OR REPLACE INTO config (key, value) VALUES ("event_banner", ?)', (url,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'banner_url': url})

@app.route('/api/config/logo', methods=['POST'])
def upload_org_logo():
    """Organisation logo — shown in the header of every page. Mirrors the
    banner uploader; stored under the config key `org_logo`."""
    if 'logo' not in request.files:
        return jsonify({'error': 'No file'}), 400
    file = request.files['logo']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'png'
    fname = f"org_logo_{uuid.uuid4().hex[:8]}.{ext}"
    fpath = os.path.join(UPLOAD_FOLDER, fname)
    file.save(fpath)
    url = f"/uploads/{fname}"
    conn = get_db()
    conn.execute('INSERT OR REPLACE INTO config (key, value) VALUES ("org_logo", ?)', (url,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'logo_url': url})

@app.route('/api/teams/logo/<int:team_id>', methods=['POST'])
def upload_team_logo(team_id):
    if 'logo' not in request.files:
        return jsonify({'error': 'No file'}), 400
    file = request.files['logo']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'png'
    fname = f"team_logo_{team_id}_{uuid.uuid4().hex[:8]}.{ext}"
    fpath = os.path.join(UPLOAD_FOLDER, fname)
    file.save(fpath)
    url = f"/uploads/{fname}"
    conn = get_db()
    conn.execute('UPDATE teams SET logo_url = ? WHERE id = ?', (url, team_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'logo_url': url})

# ─── Team-specific view data ───
@app.route('/api/team/<int:team_id>', methods=['GET'])
def get_team_data(team_id):
    conn = get_db()
    team = conn.execute('SELECT * FROM teams WHERE id=?', (team_id,)).fetchone()
    if not team:
        conn.close()
        return jsonify({'error': 'Team not found'}), 404
    rules = conn.execute('SELECT * FROM category_rules').fetchall()
    rows = conn.execute('SELECT * FROM config').fetchall()
    config = {r['key']: r['value'] for r in rows}
    players = conn.execute('SELECT * FROM players WHERE team_id=?', (team_id,)).fetchall()
    td = format_team_metrics(team, players, rules, config)

    # Current auction state
    state = {r['key']: r['value'] for r in conn.execute('SELECT * FROM auction_state').fetchall()}
    td['auction_state'] = state

    # All teams summary with metrics
    all_teams = conn.execute('SELECT * FROM teams').fetchall()
    all_teams_result = []
    for t in all_teams:
        t_players = conn.execute('SELECT id, name, category, sold_price, photo_url FROM players WHERE team_id = ?', (t['id'],)).fetchall()
        all_teams_result.append(format_team_metrics(t, t_players, rules, config))
    td['all_teams'] = all_teams_result
    td['config'] = config
    conn.close()
    return jsonify(td)

# ─── Live Spectator Data API ───
@app.route('/api/live_data', methods=['GET'])
def get_live_data():
    conn = get_db()
    config = {r['key']: r['value'] for r in conn.execute('SELECT * FROM config').fetchall()}
    state = {r['key']: r['value'] for r in conn.execute('SELECT * FROM auction_state').fetchall()}

    # Teams with Max Allowed Bid & Reserved Purse
    teams = conn.execute('SELECT * FROM teams').fetchall()
    rules = conn.execute('SELECT * FROM category_rules').fetchall()
    teams_result = []
    for t in teams:
        players = conn.execute('SELECT id, name, category, sold_price, photo_url, sold_at FROM players WHERE team_id = ? ORDER BY sold_at DESC', (t['id'],)).fetchall()
        teams_result.append(format_team_metrics(t, players, rules, config))

    if state.get('current_player'):
        p_row = conn.execute('SELECT * FROM players WHERE name = ?', (state['current_player'],)).fetchone()
        if p_row and p_row['attributes']:
            state['attributes'] = p_row['attributes']

    # Sold players
    sold = conn.execute('''
        SELECT p.id, p.name, p.category, p.sold_price, p.base_price, p.photo_url, p.attributes, p.sold_at, t.name as team_name, t.color as team_color, t.logo_url as team_logo
        FROM players p JOIN teams t ON p.team_id = t.id
        WHERE p.status = 'sold'
        ORDER BY p.sold_at DESC
    ''').fetchall()
    
    # Unsold + passed players (both available for re-auction)
    unsold = conn.execute('''
        SELECT id, name, category, base_price, photo_url, attributes, status
        FROM players WHERE status IN ('unsold', 'passed')
        ORDER BY status ASC, name ASC
    ''').fetchall()
    
    # Stats & Top Highlights
    total_count = conn.execute('SELECT COUNT(*) as c FROM players').fetchone()['c']
    passed_count = conn.execute("SELECT COUNT(*) as c FROM players WHERE status='passed'").fetchone()['c']
    total_spent = conn.execute('SELECT COALESCE(SUM(sold_price), 0) as s FROM players WHERE status="sold"').fetchone()['s']
    top_buys = conn.execute('''
        SELECT p.name, p.sold_price, p.category, p.photo_url, p.attributes, t.name as team_name, t.color as team_color, t.logo_url as team_logo
        FROM players p JOIN teams t ON p.team_id = t.id
        WHERE p.status = 'sold'
        ORDER BY p.sold_price DESC LIMIT 5
    ''').fetchall()
    
    # Category spending analytics
    cats = conn.execute('''
        SELECT category, 
               COUNT(*) as total, 
               SUM(CASE WHEN status="sold" THEN 1 ELSE 0 END) as sold_count,
               COALESCE(SUM(CASE WHEN status="sold" THEN sold_price ELSE 0 END), 0) as total_spent,
               COALESCE(AVG(CASE WHEN status="sold" THEN sold_price ELSE NULL END), 0) as avg_price
        FROM players 
        WHERE category IS NOT NULL AND category != ''
        GROUP BY category
    ''').fetchall()
    
    cat_analytics = []
    for c in cats:
        cd = dict(c)
        cd['percent_of_total'] = round((cd['total_spent'] / total_spent * 100), 1) if total_spent > 0 else 0
        cd['avg_price'] = round(cd['avg_price'], 1)
        cat_analytics.append(cd)
    
    def parse_p(p):
        d = dict(p)
        if d.get('attributes'):
            try:
                d['attributes'] = json.loads(d['attributes']) if isinstance(d['attributes'], str) else d['attributes']
            except Exception:
                d['attributes'] = {}
        else:
            d['attributes'] = {}
        return d

    conn.close()
    return jsonify({
        'config': config,
        'auction_state': state,
        # Category quotas — the live display uses these to warn when a team has
        # already filled its maximum slots for the player's category.
        'category_rules': [dict(r) for r in rules],
        'teams': teams_result,
        'sold_players': [parse_p(p) for p in sold],
        'unsold_players': [parse_p(p) for p in unsold],
        'stats': {
            'total': total_count,
            'sold': len(sold),
            'unsold': len(unsold),
            'passed': passed_count,
            'spent': total_spent,
            'top_buys': [dict(b) for b in top_buys],
            'categories': cat_analytics
        }
    })

# ─── Set Common Base Price for All Unsold Players ───
@app.route('/api/players/set_common_base_price', methods=['POST'])
def set_common_base_price():
    data = request.json or {}
    price = float(data.get('base_price', 50))
    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE players SET base_price = ? WHERE status = "unsold"', (price,))
    c.execute('INSERT OR REPLACE INTO config (key, value) VALUES ("common_base_price", ?)', (str(price),))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'common_base_price': price})

# ─── Export CSV ───
@app.route('/api/export/csv', methods=['GET'])
def export_csv():
    conn = get_db()
    players = conn.execute('''
        SELECT p.name, p.category, p.status, p.base_price, p.sold_price, COALESCE(t.name, 'Unsold') as team_name, p.sold_at
        FROM players p LEFT JOIN teams t ON p.team_id = t.id
        ORDER BY p.status DESC, p.sold_price DESC
    ''').fetchall()
    conn.close()
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Player Name', 'Category', 'Status', 'Base Price (Lakhs)', 'Sold Price (Lakhs)', 'Team', 'Sold At'])
    for p in players:
        writer.writerow([p['name'], p['category'], p['status'], p['base_price'], p['sold_price'] or '', p['team_name'], p['sold_at'] or ''])
    
    output.seek(0)
    return app.response_class(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment;filename=auction_summary.csv'}
    )






@app.route('/api/file/smart_analyze', methods=['POST'])
def smart_analyze():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    num_teams  = max(1, int(request.form.get('num_teams', 4)))
    num_splits = max(2, min(4, int(request.form.get('num_splits', 3))))
    base_price_val = float(request.form.get('base_price', 10.0))

    file = request.files['file']
    try:
        headers, dict_rows, _ = read_raw_auction_file(file)
    except Exception as e:
        return jsonify({'error': 'Failed to parse file: ' + str(e)}), 400

    if not dict_rows:
        return jsonify({'error': 'No valid rows found in file'}), 400

    total = len(dict_rows)
    col_details = inspect_file_data(headers, dict_rows)

    age_col    = next((h for h in headers if col_details[h].get('is_age_candidate')), None)
    gender_col = next((h for h in headers if col_details[h].get('is_gender_candidate')), None)
    name_col   = next((h for h in headers if col_details[h].get('is_name_candidate')), headers[0])

    player_data = []
    for row in dict_rows:
        gender = str(row.get(gender_col) or '').strip().title() if gender_col else None
        age = None
        if age_col and row.get(age_col) is not None:
            try:
                age = int(float(row[age_col]))
            except Exception:
                pass
        pname = str(row.get(name_col) or '').strip()
        player_data.append({'gender': gender or None, 'age': age, 'name': pname})

    # name -> assigned category label, persisted to the DB so player records match
    # the quota rules (this is what makes squad/quota counting work during the auction)
    player_labels = {}

    def find_best_splits(ages, n_splits, n_teams):
        from itertools import combinations
        ages_sorted = sorted(ages)
        L = len(ages_sorted)
        split_options = [n_splits]
        for alt in [n_splits - 1, n_splits + 1]:
            if 2 <= alt <= 4 and alt not in split_options and L >= alt * n_teams:
                split_options.append(alt)

        best_tuple = None
        best_thr = []
        for s_count in split_options:
            if L < s_count * n_teams:
                continue
            n_thr = s_count - 1
            unique = sorted(set(ages_sorted))
            if len(unique) < n_thr + 1:
                continue
            for combo in combinations(range(1, len(unique)), n_thr):
                thresholds = [unique[i] for i in combo]
                boundaries = [None] + thresholds + [None]
                groups = []
                for k in range(len(boundaries) - 1):
                    lo, hi = boundaries[k], boundaries[k + 1]
                    if lo is None:
                        groups.append(sum(1 for a in ages_sorted if a < hi))
                    elif hi is None:
                        groups.append(sum(1 for a in ages_sorted if a >= lo))
                    else:
                        groups.append(sum(1 for a in ages_sorted if lo <= a < hi))
                if any(g < n_teams for g in groups):
                    continue
                remainder_sum = sum(g % n_teams for g in groups)
                unequal_count = sum(1 for g in groups if g % n_teams != 0)
                imbalance = max(groups) - min(groups)
                is_req = 1 if s_count == n_splits else 0
                # Strict prioritization: 1. remainder_sum, 2. unequal_count, 3. requested splits, 4. imbalance
                score_key = (remainder_sum, unequal_count, -is_req, imbalance)
                if best_tuple is None or score_key < best_tuple:
                    best_tuple = score_key
                    best_thr = thresholds
        return list(best_thr) if best_thr else []

    GENDER_COLORS = {
        'Male':   ['#3b82f6', '#60a5fa', '#93c5fd', '#bfdbfe'],
        'Female': ['#ec4899', '#f472b6', '#f9a8d4', '#fce7f3'],
        'Other':  ['#8b5cf6', '#a78bfa', '#c4b5fd', '#ddd6fe'],
    }
    DEFAULT_COLORS = ['#f59e0b', '#10b981', '#ef4444', '#06b6d4', '#f97316']

    def make_suggestions(ages_list, label_prefix, colors, thresholds):
        import math
        results = []
        
        def create_entry(label, cnt, color_idx):
            if cnt == 0:
                return None
            remainder = cnt % num_teams
            min_t = cnt // num_teams
            # If there is a remainder, allow 1 flex slot (min_t + 1)
            max_t = (min_t + 1) if remainder > 0 else min_t
            is_exact = (remainder == 0)
            
            if is_exact:
                desc = f"{cnt} players ÷ {num_teams} teams = Exactly {min_t} per team"
            else:
                desc = f"{cnt} players ÷ {num_teams} teams = {min_t} required (+{remainder} flex slots across teams)"
            
            return {
                'category': label,
                'count': cnt,
                'is_exact': is_exact,
                'remainder': remainder,
                'per_team_min': min_t,
                'per_team_max': max_t,
                'base_price': base_price_val,
                'color': colors[color_idx % len(colors)],
                'description': desc,
                'auto': True,
            }

        if thresholds:
            boundaries = [None] + thresholds + [None]
            for i in range(len(boundaries) - 1):
                lo, hi = boundaries[i], boundaries[i + 1]
                if lo is None:
                    label = label_prefix + ' (Under ' + str(hi) + ')'
                    cnt = sum(1 for a in ages_list if a < hi)
                elif hi is None:
                    label = label_prefix + ' (' + str(lo) + '+)'
                    cnt = sum(1 for a in ages_list if a >= lo)
                else:
                    label = label_prefix + ' (' + str(lo) + u'–' + str(hi - 1) + ')'
                    cnt = sum(1 for a in ages_list if lo <= a < hi)
                
                entry = create_entry(label, cnt, i)
                if entry: results.append(entry)
        else:
            cnt = len(ages_list) if ages_list else 0
            entry = create_entry(label_prefix, cnt, 0)
            if entry: results.append(entry)
            
        return results

    from collections import Counter, OrderedDict
    suggestions = []

    # ── Generic multi-column split: the organiser chooses which column(s) to
    # divide players by (gender, age, skill level, role, …). Numeric columns
    # (e.g. age) are binned; categorical columns use their distinct values.
    # Multiple columns produce the cross-product (e.g. "Male · 18–26"). ──
    split_by = []
    try:
        split_by = json.loads(request.form.get('split_by', '[]')) or []
    except Exception:
        split_by = []
    split_by = [c for c in split_by if c in headers]

    if split_by:
        try:
            bins_map = json.loads(request.form.get('bins', '{}')) or {}
        except Exception:
            bins_map = {}

        col_thresholds, col_is_numeric = {}, {}
        for col in split_by:
            det = col_details.get(col, {})
            vals_num, numeric_ok = [], True
            for row in dict_rows:
                v = row.get(col)
                if v is None or str(v).strip() == '':
                    continue
                try:
                    vals_num.append(int(float(v)))
                except Exception:
                    numeric_ok = False
                    break
            # Treat as numeric (bin it) only when it's an age-like / continuous column
            treat_numeric = numeric_ok and vals_num and (det.get('is_age_candidate') or len(set(vals_num)) > 6)
            if treat_numeric:
                nb = max(2, min(4, int(bins_map.get(col, num_splits))))
                col_thresholds[col] = find_best_splits(vals_num, nb, num_teams) if len(vals_num) >= nb * num_teams else find_best_splits(vals_num, nb, 1)
                col_is_numeric[col] = True
            else:
                col_thresholds[col] = None
                col_is_numeric[col] = False

        def bucket_label(col, raw):
            if raw is None or str(raw).strip() == '':
                return 'N/A'
            if col_is_numeric[col]:
                try:
                    a = int(float(raw))
                except Exception:
                    return 'N/A'
                boundaries = [None] + list(col_thresholds[col] or []) + [None]
                for i in range(len(boundaries) - 1):
                    lo, hi = boundaries[i], boundaries[i + 1]
                    if (lo is None or a >= lo) and (hi is None or a < hi):
                        if lo is None:   return 'Under ' + str(hi)
                        if hi is None:   return str(lo) + '+'
                        return str(lo) + u'–' + str(hi - 1)
                return str(a)
            return str(raw).strip().title()

        groups = OrderedDict()
        for row in dict_rows:
            key = u' · '.join(bucket_label(col, row.get(col)) for col in split_by)
            groups[key] = groups.get(key, 0) + 1
            nm = str(row.get(name_col) or '').strip()
            if nm:
                player_labels[nm] = key

        # Order groups sensibly: known skill/level words in logical order, age
        # ranges low→high, else alphabetical — so the stage never shows a jumbled
        # "Intermediate, Beginner, Advanced".
        RANK = {w: i for i, w in enumerate([
            'novice', 'beginner', 'basic', 'amateur', 'rookie', 'intermediate',
            'medium', 'average', 'advanced', 'expert', 'pro', 'professional', 'elite', 'master'])}
        def order_key(label):
            parts = label.split(u' · ')
            key = []
            for p in parts:
                pl = p.strip().lower()
                if pl in RANK:
                    key.append((0, RANK[pl]))
                elif pl.startswith('under '):
                    key.append((1, -1))
                elif pl.endswith('+'):
                    key.append((1, 9999))
                elif pl[:2].isdigit():
                    key.append((1, int(''.join(ch for ch in pl.split(u'–')[0] if ch.isdigit()) or 0)))
                else:
                    key.append((2, p.lower()))
            return key
        groups = OrderedDict(sorted(groups.items(), key=lambda kv: order_key(kv[0])))

        PALETTE = ['#f59e0b', '#10b981', '#3b82f6', '#ec4899', '#8b5cf6', '#ef4444', '#06b6d4', '#f97316', '#84cc16', '#a855f7']
        for idx, (label, cnt) in enumerate(groups.items()):
            remainder = cnt % num_teams
            min_t = cnt // num_teams
            max_t = (min_t + 1) if remainder > 0 else min_t
            is_exact = (remainder == 0)
            desc = (f"{cnt} players ÷ {num_teams} teams = Exactly {min_t} per team" if is_exact
                    else f"{cnt} players ÷ {num_teams} teams = {min_t} required (+{remainder} flex slots across teams)")
            suggestions.append({
                'category': label, 'count': cnt, 'is_exact': is_exact,
                'remainder': remainder, 'per_team_min': min_t, 'per_team_max': max_t,
                'base_price': base_price_val, 'color': PALETTE[idx % len(PALETTE)],
                'description': desc, 'auto': True,
            })
        gender_counts = {}  # skip the auto gender/age branch below
    else:
        gender_counts = Counter(p['gender'] for p in player_data if p['gender'])

    def age_label(age, thresholds, prefix):
        if not thresholds or age is None:
            return prefix
        boundaries = [None] + list(thresholds) + [None]
        for i in range(len(boundaries) - 1):
            lo, hi = boundaries[i], boundaries[i + 1]
            if (lo is None or age >= lo) and (hi is None or age < hi):
                if lo is None:  return prefix + ' (Under ' + str(hi) + ')'
                if hi is None:  return prefix + ' (' + str(lo) + '+)'
                return prefix + ' (' + str(lo) + u'–' + str(hi - 1) + ')'
        return prefix

    if gender_counts:
        for gender, g_count in sorted(gender_counts.items(), key=lambda x: -x[1]):
            ages_g = [p['age'] for p in player_data if p['gender'] == gender and p['age'] is not None]
            colors = GENDER_COLORS.get(gender, DEFAULT_COLORS)
            thresholds = find_best_splits(ages_g, num_splits, num_teams) if len(ages_g) >= num_splits * num_teams else []
            for p in player_data:
                if p['gender'] == gender and p['name']:
                    player_labels[p['name']] = age_label(p['age'], thresholds, gender) if ages_g else gender
            if ages_g:
                suggestions.extend(make_suggestions(ages_g, gender, colors, thresholds))
            else:
                per_team = g_count // num_teams
                rem = g_count % num_teams
                is_exact = (rem == 0)
                max_t = (per_team + 1) if rem > 0 else per_team
                suggestions.append({
                    'category': gender, 'count': g_count,
                    'is_exact': is_exact, 'remainder': rem,
                    'per_team_min': per_team, 'per_team_max': max_t,
                    'base_price': base_price_val, 'color': colors[0],
                    'description': f"{g_count} players ÷ {num_teams} teams = {'Exactly ' + str(per_team) if is_exact else str(per_team) + ' required (+' + str(rem) + ' flex)'}/team",
                    'auto': True,
                })
        no_gender = [p for p in player_data if not p['gender']]
        if no_gender:
            ng_ages = [p['age'] for p in no_gender if p['age'] is not None]
            thresholds = find_best_splits(ng_ages, num_splits, num_teams) if len(ng_ages) >= num_splits * num_teams else []
            for p in no_gender:
                if p['name']:
                    player_labels[p['name']] = age_label(p['age'], thresholds, 'Open')
            suggestions.extend(make_suggestions(ng_ages, 'Open', DEFAULT_COLORS, thresholds))
    elif not split_by:
        all_ages = [p['age'] for p in player_data if p['age'] is not None]
        thresholds = find_best_splits(all_ages, num_splits, num_teams) if len(all_ages) >= num_splits * num_teams else []
        for p in player_data:
            if p['name']:
                player_labels[p['name']] = age_label(p['age'], thresholds, 'Players')
        suggestions.extend(make_suggestions(all_ages, 'Players', DEFAULT_COLORS, thresholds))

    # Unequal division detection & prompts
    unequal_categories = [s for s in suggestions if not s.get('is_exact', True)]
    is_perfect_division = (len(unequal_categories) == 0)
    
    unequal_prompt = None
    if not is_perfect_division:
        total_rem = sum(s.get('remainder', 0) for s in unequal_categories)
        cat_names = ", ".join([f"'{s['category']}' ({s['count']} players, +{s['remainder']} extra)" for s in unequal_categories])
        unequal_prompt = {
            'title': '⚠️ Unequal Division Alert',
            'summary': f"{len(unequal_categories)} category/categories cannot be split equally across {num_teams} teams.",
            'details': f"Categories with remainders: {cat_names}. Total unallocated flex slots: {total_rem}.",
            'recommendation': f"Allowing 1 extra player (+1 flex slot) in affected categories ensures 100% of players are sold to a team."
        }

    # Persist the computed categories onto the imported player rows so that quota
    # counting, squad fulfilment and max-bid reserves all line up during the auction.
    # Batch by label — one UPDATE per category (a handful) instead of one per
    # player (hundreds). Hundreds of round-trips to a remote DB would otherwise
    # time out the request ("Failed to fetch").
    if player_labels:
        from collections import defaultdict
        by_label = defaultdict(list)
        for nm, lbl in player_labels.items():
            by_label[lbl].append(nm)
        conn = get_db()
        cc = conn.cursor()
        for lbl, names in by_label.items():
            placeholders = ','.join(['?'] * len(names))
            cc.execute('UPDATE players SET category=? WHERE name IN (%s)' % placeholders, [lbl] + names)
        conn.commit()
        conn.close()

    photo_cols = {h for h in headers if col_details[h].get('is_photo_candidate')}
    visible_cols = [h for h in headers if h not in photo_cols]
    preview = [{k: v for k, v in row.items() if k not in photo_cols} for row in dict_rows[:6]]

    return jsonify({
        'success': True,
        'total_players': total,
        'num_teams': num_teams,
        'num_splits': num_splits,
        'is_perfect_division': is_perfect_division,
        'has_unequal': not is_perfect_division,
        'unequal_categories': unequal_categories,
        'unequal_prompt': unequal_prompt,
        'age_col': age_col,
        'gender_col': gender_col,
        'split_by': split_by,
        'suggestions': suggestions,
        'preview': preview,
        'columns': visible_cols,
    })



@app.route('/api/action/pass', methods=['POST'])
def pass_player():
    data = request.json or {}
    player_id = data.get('player_id')
    if not player_id:
        return jsonify({'error': 'Player ID required'}), 400

    conn = get_db()
    c = conn.cursor()
    player = c.execute('SELECT * FROM players WHERE id=?', (player_id,)).fetchone()
    if not player:
        conn.close()
        return jsonify({'error': 'Player not found'}), 404

    c.execute('''INSERT INTO action_history
        (action_type, player_id, player_name, old_status, new_status, base_price, category, photo_url)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
        ('pass', player_id, player['name'], player['status'], 'passed',
         player['base_price'], player['category'], player['photo_url']))

    c.execute('UPDATE players SET status="passed", team_id=NULL, sold_price=NULL WHERE id=?', (player_id,))
    # Clear current player from stage + write UNSOLD state for Cinematic Stage overlay
    c.execute("DELETE FROM auction_state WHERE key NOT IN ('auction_sport','auction_template_mode','auction_template')")
    c.execute("INSERT OR REPLACE INTO auction_state (key, value) VALUES ('auction_status', 'passed')")
    c.execute("INSERT OR REPLACE INTO auction_state (key, value) VALUES ('last_passed_player', ?)", (player['name'],))
    c.execute("INSERT OR REPLACE INTO auction_state (key, value) VALUES ('last_passed_photo', ?)", (player['photo_url'] or '',))
    conn.commit()
    conn.close()
    excel_backup_async()
    return jsonify({'success': True, 'player_name': player['name']})

@app.route('/api/action/revive', methods=['POST'])
def revive_player():
    data = request.json or {}
    player_id = data.get('player_id')
    half_price = data.get('half_price', False)
    
    if not player_id:
        return jsonify({'error': 'Player ID required'}), 400
        
    conn = get_db()
    c = conn.cursor()
    if half_price:
        c.execute('UPDATE players SET status="unsold", base_price = ROUND(base_price / 2.0, 1) WHERE id=?', (player_id,))
    else:
        c.execute('UPDATE players SET status="unsold" WHERE id=?', (player_id,))
    conn.commit()
    conn.close()
    excel_backup_async()
    return jsonify({'success': True})

@app.route('/api/action/bargain_bin', methods=['POST'])
def bargain_bin():
    conn = get_db()
    try:
        c = conn.cursor()
        unsold = c.execute("SELECT id, base_price FROM players WHERE status='unsold'").fetchall()
        count = 0
        for p in unsold:
            new_price = max(1, round(p['base_price'] / 2.0, 1))
            c.execute("UPDATE players SET base_price=? WHERE id=?", (new_price, p['id']))
            count += 1
        conn.commit()
        excel_backup_async()
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

if __name__ == '__main__':
    init_db()
    Timer(1, lambda: webbrowser.open_new('http://127.0.0.1:5000/')).start()
    # use_debugger=False disables the interactive Werkzeug pin console so crashed
    # requests don't hold SQLite write locks open; error tracebacks still appear.
    app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False,
            use_debugger=False, threaded=True)
