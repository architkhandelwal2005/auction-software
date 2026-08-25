import os
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure directories exist
os.makedirs("sample_data", exist_ok=True)

# Random Seed for Reproducibility
random.seed(42)

male_first_names = [
    "Aarav", "Rohan", "Vikram", "Siddharth", "Rahul", "Arjun", "Kabir", "Aditya", "Aryan", "Kunal",
    "Dev", "Ishan", "Amit", "Neeraj", "Rajesh", "Suresh", "Prateek", "Ankit", "Yash", "Varun",
    "Nikhil", "Gaurav", "Mayank", "Tushar", "Abhishek", "Karan", "Alok", "Hardik", "Shubman", "Rishabh",
    "Suryakumar", "Shreyas", "Lokesh", "Ravindra", "Jasprit", "Mohammed", "Axar", "Kuldeep", "Sanju", "Rinku",
    "Shivam", "Tilak", "Yashasvi", "Washington", "Avesh", "Arshdeep", "Mukesh", "Ravi", "Deepak", "Harshal",
    "Chetan", "Krunal", "Jaydev", "Vijay", "Sanjay", "Manoj", "Manish", "Piyush", "Sandeep", "Robin",
    "Dinesh", "Parthiv", "Irfan", "Yusuf", "Zaheer", "Ashish", "Harbhajan", "Pragyan", "Munaf", "Praveen",
    "Vinay", "Ashok", "Umesh", "Dhawal", "Karn", "Jayant", "Kulwant", "Aniket", "Basil", "Kamlesh",
    "Riyan", "Prabhsimran", "Shahrukh", "Sai", "Dhruv", "Atharva", "Ayush", "Sameer", "Tanush", "Harsh",
    "Priyanshu", "Vaibhav", "Manan", "Krish", "Vivaan", "Vihaan", "Reyansh", "Advik", "Samar", "Ayaan"
]

female_first_names = [
    "Ananya", "Priya", "Sneha", "Pooja", "Riya", "Neha", "Tanvi", "Shreya", "Divya", "Aditi",
    "Ishita", "Meera", "Kavya", "Ritu", "Deepa", "Swati", "Payal", "Sonam", "Kriti", "Shruti",
    "Radhika", "Mansi", "Sakshi", "Natasha", "Simran", "Nikita", "Anjali", "Jyoti", "Sunita", "Rekha",
    "Anita", "Geeta", "Preeti", "Komal", "Rashmi", "Suman", "Shilpa", "Poonam", "Archana", "Vandana",
    "Sushma", "Alka", "Sarita", "Usha", "Shanti", "Lata", "Asha", "Nirmala", "Manju", "Seema",
    "Smriti", "Harmanpreet", "Shafali", "Jemimah", "Deepti", "Richa", "Renuka", "Radha", "Rajeshwari", "Meghna",
    "Yastika", "Harleen", "Sneh", "Shikha", "Taniya", "Ekta", "Veda", "Punam", "Mona", "Arundhati",
    "Shweta", "Parshavi", "Soumya", "Titas", "Mannat", "Kanika", "Vrinda", "Kashvee", "Amanjot", "Disha",
    "Jasia", "Dhara", "Tejal", "Uma", "Sajeevan", "Minnu", "Pratika", "Tarannum", "Sayali", "Anuja",
    "Devika", "Nuzhat", "Bhavna", "Chhavi", "Tanushree", "Pallavi", "Ira", "Bratati", "Lavanya", "Mahika"
]

last_names = [
    "Sharma", "Verma", "Patel", "Mehta", "Gupta", "Singh", "Kumar", "Joshi", "Shah", "Rao",
    "Reddy", "Nair", "Das", "Sen", "Roy", "Banerjee", "Mukherjee", "Chatterjee", "Iyer", "Iyengar",
    "Pillai", "Menon", "Nambiar", "Kulkarni", "Deshmukh", "Patil", "Shinde", "Jadhav", "Pawar", "Gaikwad",
    "Chavan", "Bhosle", "More", "Salunkhe", "Sawant", "Rane", "Desai", "Trivedi", "Pandya", "Bhatt",
    "Dave", "Shukla", "Mishra", "Tiwari", "Pandey", "Tripathi", "Dubey", "Chaubey", "Upadhyay", "Dwivedi",
    "Pathak", "Awasthi", "Agrawal", "Bansal", "Mittal", "Goyal", "Garg", "Singhal", "Jindal", "Goel",
    "Kansal", "Kapoor", "Malhotra", "Khanna", "Chopra", "Bhatia", "Sethi", "Kohli", "Dhawan", "Gambhir",
    "Chawla", "Bhandari", "Sarin", "Kohli", "Thakur", "Rana", "Chauhan", "Rawat", "Bisht", "Negi",
    "Kashyap", "Saxena", "Srivastava", "Mathur", "Nigam", "Asthana", "Sinha", "Srivastava", "Prasad", "Sahu"
]

def generate_phone():
    prefix = random.choice(["98", "97", "99", "96", "95", "93", "91", "88", "87", "89", "85", "80", "79", "78", "77", "70"])
    rest = "".join([str(random.randint(0, 9)) for _ in range(8)])
    return f"{prefix}{rest}"

def get_age_cat_cricket(age):
    if age < 19:
        return "Under-19"
    elif age <= 28:
        return "Young Stars (19-28)"
    elif age <= 38:
        return "Senior Pro (29-38)"
    else:
        return "Legends (39+)"

def get_age_cat_badminton(age):
    if age < 16:
        return "Sub-Junior (U-16)"
    elif age < 20:
        return "Junior (U-20)"
    elif age <= 35:
        return "Open / Elite (20-35)"
    else:
        return "Masters (36+)"

def get_age_cat_corporate(age):
    if age <= 28:
        return "Junior Associate (<=28)"
    elif age <= 38:
        return "Mid-Management (29-38)"
    elif age <= 48:
        return "Senior Leadership (39-48)"
    else:
        return "Executive Veteran (49+)"

def get_age_cat_society(age):
    if age <= 15:
        return "Society Kids (U-15)"
    elif age <= 25:
        return "Youth Wing (16-25)"
    elif age <= 45:
        return "Prime Adults (26-45)"
    else:
        return "Senior Citizens (46+)"

def get_age_cat_youth(age):
    if age <= 12:
        return "Sub-Junior U-12"
    elif age <= 15:
        return "Junior U-15"
    elif age <= 18:
        return "Youth U-18"
    else:
        return "Emerging Pro (19-21)"

def style_worksheet(ws, header_fill_hex="1E3A8A"):
    # Header styling
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Zebra striping
    row_fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    ws.row_dimensions[1].height = 28
    
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        fill = row_fill_even if is_even else row_fill_odd
        for cell in ws[row_idx]:
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(name="Segoe UI", size=10)
            if isinstance(cell.value, (int, float)):
                cell.alignment = center_align
            elif str(cell.value).startswith("+91") or str(cell.value).isdigit():
                cell.alignment = center_align
            elif cell.value in ["Male", "Female", "M", "F", "Other"]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)


# ==========================================
# FILE 1: Cricket Premier League (100 Players)
# Ratio: 60 Male / 40 Female | Age: 15 to 48
# ==========================================
def create_cricket_file(filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cricket Players"
    
    headers = ["Player ID", "Full Name", "Gender", "Age", "Age Category", "Primary Role", "Batting Hand", "Bowling Style", "Mobile Number", "Base Price"]
    ws.append(headers)
    
    roles = [
        ("Top-order Batsman", "Right-Handed", "Right-arm Medium"),
        ("Middle-order Batsman", "Right-Handed", "Right-arm Offbreak"),
        ("Opening Batsman", "Left-Handed", "Slow Left-arm"),
        ("All-Rounder", "Right-Handed", "Right-arm Fast"),
        ("All-Rounder", "Left-Handed", "Slow Left-arm Orthodox"),
        ("Wicketkeeper-Batsman", "Right-Handed", "None"),
        ("Fast Bowler", "Right-Handed", "Right-arm Fast-Medium"),
        ("Spin Bowler", "Right-Handed", "Right-arm Legbreak"),
        ("Pace Bowler", "Left-Handed", "Left-arm Fast")
    ]
    
    # 60 Male, 40 Female
    genders = ["Male"] * 60 + ["Female"] * 40
    random.shuffle(genders)
    
    used_names = set()
    for i, gender in enumerate(genders, 1):
        pid = f"CPL-{i:03d}"
        fnames = male_first_names if gender == "Male" else female_first_names
        
        while True:
            fn = random.choice(fnames)
            ln = random.choice(last_names)
            name = f"{fn} {ln}"
            if name not in used_names:
                used_names.add(name)
                break
                
        r = random.random()
        if r < 0.20:
            age = random.randint(15, 18) # U-19
        elif r < 0.65:
            age = random.randint(19, 28) # Young stars
        elif r < 0.90:
            age = random.randint(29, 38) # Senior pro
        else:
            age = random.randint(39, 48) # Legends
            
        age_cat = get_age_cat_cricket(age)
        role, bat, bowl = random.choice(roles)
        phone = generate_phone()
        base_price = random.choice([20, 30, 40, 50, 60, 75, 100, 150])
        
        ws.append([pid, name, gender, age, age_cat, role, bat, bowl, phone, base_price])
        
    style_worksheet(ws, "1E40AF") # Royal Blue
    wb.save(filepath)
    print(f"Created: {filepath} ({len(genders)} players)")


# ==========================================
# FILE 2: Badminton Club Tournament (95 Players)
# Ratio: 45 Male / 50 Female (Female-Heavy) | Age: 11 to 55
# ==========================================
def create_badminton_file(filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Badminton Roster"
    
    headers = ["Reg No", "Player Name", "Gender", "Age", "Age Bracket", "Playing Category", "Dominant Hand", "Contact Number", "Base Price"]
    ws.append(headers)
    
    categories = ["Singles Specialist", "Doubles Specialist", "Mixed Doubles Specialist", "All-Round Player"]
    dominant_hands = ["Right", "Right", "Right", "Left"]
    
    genders = ["Male"] * 45 + ["Female"] * 50
    random.shuffle(genders)
    
    used_names = set()
    for i, gender in enumerate(genders, 1):
        reg_no = f"BC-2026-{i:03d}"
        fnames = male_first_names if gender == "Male" else female_first_names
        
        while True:
            fn = random.choice(fnames)
            ln = random.choice(last_names)
            name = f"{fn} {ln}"
            if name not in used_names:
                used_names.add(name)
                break
                
        r = random.random()
        if r < 0.22:
            age = random.randint(11, 15) # Sub-Junior
        elif r < 0.48:
            age = random.randint(16, 19) # Junior
        elif r < 0.82:
            age = random.randint(20, 35) # Open / Elite
        else:
            age = random.randint(36, 56) # Masters
            
        age_bracket = get_age_cat_badminton(age)
        play_cat = random.choice(categories)
        hand = random.choice(dominant_hands)
        phone = generate_phone()
        base_price = random.choice([25, 50, 75, 100, 125, 150, 200])
        
        ws.append([reg_no, name, gender, age, age_bracket, play_cat, hand, phone, base_price])
        
    style_worksheet(ws, "0F766E") # Teal / Emerald
    wb.save(filepath)
    print(f"Created: {filepath} ({len(genders)} players)")


# ==========================================
# FILE 3: Corporate Sports Olympiad (90 Players)
# Ratio: 58 Male / 32 Female | Age: 22 to 54
# ==========================================
def create_corporate_file(filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corporate Athletes"
    
    headers = ["Emp Code", "Employee Name", "Gender", "Age", "Age Division", "Department", "Primary Sport", "Secondary Sport", "Mobile Number", "Base Price"]
    ws.append(headers)
    
    departments = ["Engineering & Tech", "Sales & Marketing", "Finance & Accounts", "Human Resources", "Operations & Supply", "Product & Design", "Legal & Compliance"]
    sports = ["Table Tennis", "Football", "Badminton", "Cricket", "Volleyball", "Chess", "Basketball", "Carrom"]
    
    genders = ["Male"] * 58 + ["Female"] * 32
    random.shuffle(genders)
    
    used_names = set()
    for i, gender in enumerate(genders, 1):
        emp_code = f"CORP-{1000 + i}"
        fnames = male_first_names if gender == "Male" else female_first_names
        
        while True:
            fn = random.choice(fnames)
            ln = random.choice(last_names)
            name = f"{fn} {ln}"
            if name not in used_names:
                used_names.add(name)
                break
                
        r = random.random()
        if r < 0.35:
            age = random.randint(22, 28)
        elif r < 0.70:
            age = random.randint(29, 38)
        elif r < 0.90:
            age = random.randint(39, 48)
        else:
            age = random.randint(49, 56)
            
        age_div = get_age_cat_corporate(age)
        dept = random.choice(departments)
        p_sport = random.choice(sports)
        s_sports = [s for s in sports if s != p_sport]
        s_sport = random.choice(s_sports)
        phone = generate_phone()
        base_price = random.choice([50, 75, 100, 120, 150, 200])
        
        ws.append([emp_code, name, gender, age, age_div, dept, p_sport, s_sport, phone, base_price])
        
    style_worksheet(ws, "4338CA") # Indigo
    wb.save(filepath)
    print(f"Created: {filepath} ({len(genders)} players)")


# ==========================================
# FILE 4: Residential Township Mega League (115 Players)
# Ratio: 65 Male / 50 Female | Age: 9 to 68
# ==========================================
def create_society_file(filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Society Residents"
    
    headers = ["Resident ID", "Full Name", "Gender", "Age", "Age Classification", "Wing / Flat No", "Preferred Sport", "Skill Level", "Contact Number", "Base Price"]
    ws.append(headers)
    
    wings = ["A", "B", "C", "D", "E", "F", "Tower 1", "Tower 2", "Tower 3"]
    sports = ["Box Cricket", "Badminton", "Table Tennis", "Pickleball", "Carrom & Chess", "Swimming", "Lawn Tennis"]
    skill_levels = ["Beginner", "Intermediate", "Advanced", "Pro / Semi-Pro"]
    
    genders = ["Male"] * 65 + ["Female"] * 50
    random.shuffle(genders)
    
    used_names = set()
    for i, gender in enumerate(genders, 1):
        rid = f"RES-{i:03d}"
        fnames = male_first_names if gender == "Male" else female_first_names
        
        while True:
            fn = random.choice(fnames)
            ln = random.choice(last_names)
            name = f"{fn} {ln}"
            if name not in used_names:
                used_names.add(name)
                break
                
        r = random.random()
        if r < 0.22:
            age = random.randint(9, 15)   # Kids
        elif r < 0.50:
            age = random.randint(16, 25)  # Youth
        elif r < 0.82:
            age = random.randint(26, 45)  # Adults
        else:
            age = random.randint(46, 68)  # Seniors
            
        age_class = get_age_cat_society(age)
        flat = f"{random.choice(wings)}-{random.randint(1, 19):02d}{random.randint(1, 4):02d}"
        sport = random.choice(sports)
        skill = random.choice(skill_levels)
        phone = generate_phone()
        base_price = random.choice([10, 20, 30, 50, 75, 100])
        
        ws.append([rid, name, gender, age, age_class, flat, sport, skill, phone, base_price])
        
    style_worksheet(ws, "7C3AED") # Purple
    wb.save(filepath)
    print(f"Created: {filepath} ({len(genders)} players)")


# ==========================================
# FILE 5: Youth Sports & Academy Hunt (90 Players)
# Ratio: 48 Male / 42 Female | Age: 8 to 21
# ==========================================
def create_youth_academy_file(filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Academy Athletes"
    
    headers = ["Academy ID", "Athlete Name", "Gender", "Age", "Age Group", "Discipline", "State / Region", "Parent Contact", "Base Price"]
    ws.append(headers)
    
    disciplines = ["Athletics", "Football", "Cricket", "Basketball", "Swimming", "Badminton", "Gymnastics", "Martial Arts"]
    states = ["Maharashtra", "Karnataka", "Delhi NCR", "Tamil Nadu", "Punjab", "Haryana", "Gujarat", "West Bengal", "Telangana", "Kerala"]
    
    genders = ["Male"] * 48 + ["Female"] * 42
    random.shuffle(genders)
    
    used_names = set()
    for i, gender in enumerate(genders, 1):
        aid = f"ACA-2026-{i:03d}"
        fnames = male_first_names if gender == "Male" else female_first_names
        
        while True:
            fn = random.choice(fnames)
            ln = random.choice(last_names)
            name = f"{fn} {ln}"
            if name not in used_names:
                used_names.add(name)
                break
                
        r = random.random()
        if r < 0.28:
            age = random.randint(8, 12)  # Sub-Junior U-12
        elif r < 0.65:
            age = random.randint(13, 15) # Junior U-15
        elif r < 0.90:
            age = random.randint(16, 18) # Youth U-18
        else:
            age = random.randint(19, 21) # Emerging Pro
            
        age_group = get_age_cat_youth(age)
        discipline = random.choice(disciplines)
        state = random.choice(states)
        phone = generate_phone()
        base_val = random.choice([20, 30, 40, 50, 70, 90, 100])
        
        ws.append([aid, name, gender, age, age_group, discipline, state, phone, base_val])
        
    style_worksheet(ws, "EA580C") # Orange
    wb.save(filepath)
    print(f"Created: {filepath} ({len(genders)} players)")


if __name__ == "__main__":
    files = [
        ("sample_cricket_tournament_100_players.xlsx", create_cricket_file),
        ("sample_badminton_championship_95_players.xlsx", create_badminton_file),
        ("sample_corporate_sports_meet_90_players.xlsx", create_corporate_file),
        ("sample_society_mega_league_115_players.xlsx", create_society_file),
        ("sample_youth_academy_hunt_90_players.xlsx", create_youth_academy_file),
    ]
    
    # Save both in root and in sample_data
    for fname, generator_func in files:
        root_path = fname
        sample_path = os.path.join("sample_data", fname)
        generator_func(root_path)
        import shutil
        shutil.copyfile(root_path, sample_path)
        print(f"Copied to {sample_path}")

    print("\nAll 5 sample Excel files successfully generated!")
